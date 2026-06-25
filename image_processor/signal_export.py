"""Signal extraction and document export for Femtonics multi-ROI TIFF stacks."""

from __future__ import annotations

import csv
from dataclasses import dataclass
from datetime import datetime
from pathlib import Path
from typing import Callable, Iterable

import numpy as np
from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter
from PIL import Image

try:
    from .roi_core import (
        RoiLayout,
        RoiMetadata,
        TiffInfo,
        compute_roi_layout,
        load_tiff_info,
        metadata_path_for_tiff,
        parse_metadata,
        validate_tiff_matches_metadata,
    )
    from .tiff_backend import iter_tiff_frames
except ImportError:  # pragma: no cover - supports direct script-style imports.
    from roi_core import (
        RoiLayout,
        RoiMetadata,
        TiffInfo,
        compute_roi_layout,
        load_tiff_info,
        metadata_path_for_tiff,
        parse_metadata,
        validate_tiff_matches_metadata,
    )
    from tiff_backend import iter_tiff_frames


SignalProgress = Callable[[int, int], None]


@dataclass(frozen=True)
class SignalTable:
    source_tiff: Path
    metadata_path: Path
    metadata: RoiMetadata
    tiff_info: TiffInfo
    layout: RoiLayout
    sampling_rate_hz: float
    time_ms: np.ndarray
    signals: np.ndarray
    processed_at: str

    @property
    def roi_labels(self) -> list[str]:
        return [f"ROI{box.roi_index:02d}" for box in self.layout.boxes]

    @property
    def headers(self) -> list[str]:
        return ["Frame", "Time_ms", *self.roi_labels]

    def iter_signal_rows(self) -> Iterable[list[float | int]]:
        for frame_index in range(self.tiff_info.frame_count):
            yield [
                frame_index,
                float(self.time_ms[frame_index]),
                *[float(value) for value in self.signals[frame_index, :]],
            ]


@dataclass(frozen=True)
class ExportResult:
    signal_table: SignalTable
    xlsx_path: Path
    csv_path: Path | None


def resolve_sampling_rate(metadata: RoiMetadata, sampling_rate_hz: float | None = None) -> float:
    rate = sampling_rate_hz if sampling_rate_hz is not None else metadata.sampling_rate_hz
    if rate is None:
        raise ValueError("Sampling rate is required for Time_ms export.")
    rate = float(rate)
    if rate <= 0:
        raise ValueError(f"Sampling rate must be greater than zero, got {rate}.")
    return rate


def _frame_roi_means(frame: np.ndarray, layout: RoiLayout) -> np.ndarray:
    values = np.asarray(frame, dtype=np.float64)
    if (
        values.ndim == 2
        and values.shape == (layout.rows * layout.roi_height, layout.columns * layout.roi_width)
    ):
        grid = values.reshape(layout.rows, layout.roi_height, layout.columns, layout.roi_width)
        return np.asarray(grid.mean(axis=(1, 3)).reshape(-1), dtype=np.float64)
    return np.asarray(
        [
            float(np.mean(values[box.upper : box.lower, box.left : box.right]))
            for box in layout.boxes
        ],
        dtype=np.float64,
    )


def extract_roi_signals(
    tiff_path: Path,
    sampling_rate_hz: float | None = None,
    progress: SignalProgress | None = None,
) -> SignalTable:
    metadata_path = metadata_path_for_tiff(tiff_path)
    if not metadata_path.exists():
        raise FileNotFoundError(f"Metadata file is missing: {metadata_path}")

    metadata = parse_metadata(metadata_path)
    rate = resolve_sampling_rate(metadata, sampling_rate_hz)
    layout = compute_roi_layout(metadata)
    tiff_info = load_tiff_info(tiff_path)
    validate_tiff_matches_metadata(tiff_info, metadata)

    signals = np.zeros((tiff_info.frame_count, len(layout.boxes)), dtype=np.float64)
    for frame_number, frame in enumerate(iter_tiff_frames(tiff_path)):
        if frame_number >= tiff_info.frame_count:
            break
        signals[frame_number, :] = _frame_roi_means(frame, layout)

        if progress is not None and (
            frame_number == 0
            or (frame_number + 1) % 500 == 0
            or frame_number + 1 == tiff_info.frame_count
        ):
            progress(frame_number + 1, tiff_info.frame_count)

    frame_index = np.arange(tiff_info.frame_count, dtype=np.float64)
    time_ms = frame_index * 1000.0 / rate
    return SignalTable(
        source_tiff=tiff_path,
        metadata_path=metadata_path,
        metadata=metadata,
        tiff_info=tiff_info,
        layout=layout,
        sampling_rate_hz=rate,
        time_ms=time_ms,
        signals=signals,
        processed_at=datetime.now().isoformat(timespec="seconds"),
    )


def default_export_output_dir(tiff_path: Path, output_dir: Path | None = None) -> Path:
    if output_dir is not None:
        return output_dir
    return tiff_path.parent / "exports"


def write_signals_csv(signal_table: SignalTable, output_path: Path) -> None:
    output_path.parent.mkdir(parents=True, exist_ok=True)
    with output_path.open("w", newline="", encoding="utf-8") as handle:
        writer = csv.writer(handle)
        writer.writerow(signal_table.headers)
        writer.writerows(signal_table.iter_signal_rows())


def metadata_sheet_rows(signal_table: SignalTable) -> list[tuple[str, str | int | float | None]]:
    metadata = signal_table.metadata
    info = signal_table.tiff_info
    layout = signal_table.layout
    return [
        ("Source TIFF", str(signal_table.source_tiff)),
        ("Metadata file", str(signal_table.metadata_path)),
        ("Processed at", signal_table.processed_at),
        ("Signal calculation", "Mean raw pixel intensity inside each rectangular ROI"),
        ("Sampling rate Hz", signal_table.sampling_rate_hz),
        ("Time unit", "ms"),
        ("Frame count", info.frame_count),
        ("Frame width", info.frame_width),
        ("Frame height", info.frame_height),
        ("TIFF mode", info.mode),
        ("TIFF dtype", info.dtype),
        ("Acquisition type", metadata.acquisition_type),
        ("Comment", metadata.comment),
        ("Measurement date", metadata.measurement_date),
        ("Channel", metadata.channel),
        ("Metadata data type", metadata.data_type),
        ("Average", metadata.average),
        ("Pattern number", metadata.pattern_num),
        ("ROI count", metadata.roi_count),
        ("ROI order", layout.order),
        ("ROI grid columns", layout.columns),
        ("ROI grid rows", layout.rows),
        ("ROI width", layout.roi_width),
        ("ROI height", layout.roi_height),
    ]


def write_signals_xlsx(signal_table: SignalTable, output_path: Path) -> None:
    output_path.parent.mkdir(parents=True, exist_ok=True)
    workbook = Workbook()
    signals_sheet = workbook.active
    signals_sheet.title = "Signals"

    header_fill = PatternFill("solid", fgColor="1F4E78")
    header_font = Font(bold=True, color="FFFFFF")
    subtle_fill = PatternFill("solid", fgColor="EAF2F8")

    signals_sheet.append(signal_table.headers)
    for row in signal_table.iter_signal_rows():
        signals_sheet.append(row)

    for cell in signals_sheet[1]:
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center")

    signals_sheet.freeze_panes = "C2"
    signals_sheet.auto_filter.ref = signals_sheet.dimensions
    signals_sheet.column_dimensions["A"].width = 12
    signals_sheet.column_dimensions["B"].width = 14
    for col_index in range(3, len(signal_table.headers) + 1):
        signals_sheet.column_dimensions[get_column_letter(col_index)].width = 13
    for row in signals_sheet.iter_rows(min_row=2, min_col=2, max_col=len(signal_table.headers)):
        for cell in row:
            cell.number_format = "0.000"

    metadata_sheet = workbook.create_sheet("Metadata")
    metadata_sheet.append(["Field", "Value"])
    for row in metadata_sheet_rows(signal_table):
        metadata_sheet.append(list(row))
    for cell in metadata_sheet[1]:
        cell.fill = header_fill
        cell.font = header_font
    metadata_sheet.column_dimensions["A"].width = 24
    metadata_sheet.column_dimensions["B"].width = 96
    metadata_sheet.freeze_panes = "A2"

    roi_sheet = workbook.create_sheet("ROI_Map")
    roi_sheet.append(
        [
            "ROI",
            "Shape",
            "Row",
            "Column",
            "Left",
            "Upper",
            "Right",
            "Lower",
            "Width",
            "Height",
        ]
    )
    for box in signal_table.layout.boxes:
        roi_sheet.append(
            [
                f"ROI{box.roi_index:02d}",
                box.shape_type,
                box.row + 1,
                box.column + 1,
                box.left,
                box.upper,
                box.right,
                box.lower,
                box.width,
                box.height,
            ]
        )
    for cell in roi_sheet[1]:
        cell.fill = header_fill
        cell.font = header_font
    for row in range(2, roi_sheet.max_row + 1):
        if row % 2 == 0:
            for cell in roi_sheet[row]:
                cell.fill = subtle_fill
    roi_sheet.freeze_panes = "A2"
    for col_index in range(1, roi_sheet.max_column + 1):
        roi_sheet.column_dimensions[get_column_letter(col_index)].width = 13

    workbook.save(output_path)


def export_signal_documents(
    tiff_path: Path,
    output_dir: Path | None = None,
    sampling_rate_hz: float | None = None,
    write_csv: bool = True,
    progress: SignalProgress | None = None,
) -> ExportResult:
    signal_table = extract_roi_signals(tiff_path, sampling_rate_hz=sampling_rate_hz, progress=progress)
    export_dir = default_export_output_dir(tiff_path, output_dir)
    export_dir.mkdir(parents=True, exist_ok=True)

    xlsx_path = export_dir / f"{tiff_path.stem}_signals.xlsx"
    csv_path = export_dir / f"{tiff_path.stem}_signals.csv" if write_csv else None

    write_signals_xlsx(signal_table, xlsx_path)
    if csv_path is not None:
        write_signals_csv(signal_table, csv_path)

    return ExportResult(signal_table=signal_table, xlsx_path=xlsx_path, csv_path=csv_path)
