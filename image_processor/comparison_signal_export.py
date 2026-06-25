"""Workbook export for native pipeline signal comparison variants."""

from __future__ import annotations

from pathlib import Path
from typing import Any

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter


VARIANT_SHEETS = [
    ("raw", "Raw"),
    ("tracked_segmentation", "Tracked Segmentation"),
    ("segmentation", "Segmented"),
    ("motion_segmentation_on_raw", "Motion Seg on Raw"),
    ("motion_corrected", "Motion Corrected"),
    ("motion_corrected_segmentation", "Segmented+Motion"),
]


def _style_header(row: Any, fill: PatternFill, font: Font) -> None:
    for cell in row:
        cell.fill = fill
        cell.font = font
        cell.alignment = Alignment(horizontal="center")


def _write_signal_sheet(sheet: Any, signal_table: Any, header_fill: PatternFill, header_font: Font) -> None:
    sheet.append(signal_table.headers)
    for row in signal_table.iter_signal_rows():
        sheet.append(row)
    _style_header(sheet[1], header_fill, header_font)
    sheet.freeze_panes = "C2"
    sheet.auto_filter.ref = sheet.dimensions
    sheet.column_dimensions["A"].width = 12
    sheet.column_dimensions["B"].width = 14
    for col_index in range(3, len(signal_table.headers) + 1):
        sheet.column_dimensions[get_column_letter(col_index)].width = 16
    for row in sheet.iter_rows(min_row=2, min_col=2, max_col=len(signal_table.headers)):
        for cell in row:
            cell.number_format = "0.000"


def _write_skipped_sheet(sheet: Any, variant: dict[str, Any], header_fill: PatternFill, header_font: Font) -> None:
    sheet.append(["Status", "Detail"])
    if variant.get("skipped"):
        status = "Skipped"
        detail = variant.get("reason", "")
    else:
        status = "Unavailable"
        detail = variant.get("error", "Signal table was not generated.")
    sheet.append([status, detail])
    _style_header(sheet[1], header_fill, header_font)
    sheet.column_dimensions["A"].width = 16
    sheet.column_dimensions["B"].width = 96


def write_signal_comparison_workbook(
    variants: dict[str, dict[str, Any]],
    signal_tables: dict[str, Any],
    output_path: Path,
) -> Path:
    """Write one workbook with stable sheets for raw/motion/segmentation comparisons."""
    output_path.parent.mkdir(parents=True, exist_ok=True)
    workbook = Workbook()
    header_fill = PatternFill("solid", fgColor="1F4E78")
    header_font = Font(bold=True, color="FFFFFF")
    subtle_fill = PatternFill("solid", fgColor="EAF2F8")

    first = True
    for variant_name, sheet_name in VARIANT_SHEETS:
        sheet = workbook.active if first else workbook.create_sheet(sheet_name)
        sheet.title = sheet_name
        first = False

        table = signal_tables.get(variant_name)
        variant = variants.get(variant_name, {})
        if table is not None and variant.get("ok") and not variant.get("skipped"):
            _write_signal_sheet(sheet, table, header_fill, header_font)
        else:
            _write_skipped_sheet(sheet, variant, header_fill, header_font)

    metadata_sheet = workbook.create_sheet("Export Map")
    metadata_sheet.append(
        [
            "Variant",
            "Sheet",
            "Status",
            "Signal_Source",
            "Source_TIFF",
            "Mask_Source_TIFF",
            "Mask_Path",
            "CSV",
            "ROI_Count",
            "Frame_Count",
            "Detail",
        ]
    )
    for variant_name, sheet_name in VARIANT_SHEETS:
        variant = variants.get(variant_name, {})
        if variant.get("ok") and not variant.get("skipped"):
            status = "OK"
            detail = ""
        elif variant.get("skipped"):
            status = "Skipped"
            detail = variant.get("reason", "")
        else:
            status = "Failed"
            detail = variant.get("error", "")
        metadata_sheet.append(
            [
                variant_name,
                sheet_name,
                status,
                variant.get("signal_source"),
                variant.get("source_tiff"),
                variant.get("mask_source_tiff"),
                variant.get("mask_path"),
                variant.get("csv"),
                variant.get("roi_count"),
                variant.get("frame_count"),
                detail,
            ]
        )
    _style_header(metadata_sheet[1], header_fill, header_font)
    for row_number in range(2, metadata_sheet.max_row + 1):
        if row_number % 2 == 0:
            for cell in metadata_sheet[row_number]:
                cell.fill = subtle_fill
    metadata_sheet.freeze_panes = "A2"
    widths = [28, 22, 14, 22, 72, 72, 72, 72, 12, 12, 80]
    for col_index, width in enumerate(widths, start=1):
        metadata_sheet.column_dimensions[get_column_letter(col_index)].width = width

    workbook.save(output_path)
    return output_path
