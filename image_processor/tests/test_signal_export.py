import csv
import json
import sys
import tempfile
import unittest
from pathlib import Path

import numpy as np
from openpyxl import load_workbook
from PIL import Image

from image_processor import export_signals, native_pipeline, validate_hooks
from image_processor.roi_core import extract_sampling_rate_hz, metadata_path_for_tiff, parse_metadata
from image_processor.signal_export import export_signal_documents, extract_roi_signals


class SignalExportTests(unittest.TestCase):
    def test_parse_sampling_rate_and_export_excel_csv(self) -> None:
        with tempfile.TemporaryDirectory() as temp_dir:
            root = Path(temp_dir)
            tiff_path = root / "sample.tiff"
            metadata_path = metadata_path_for_tiff(tiff_path)
            self._write_synthetic_tiff(tiff_path)
            metadata_path.write_text(self._metadata_text("Comment: 6rois-500hz"), encoding="utf-8")

            metadata = parse_metadata(metadata_path)
            self.assertEqual(metadata.sampling_rate_hz, 500.0)

            signal_table = extract_roi_signals(tiff_path)
            self.assertEqual(signal_table.headers, ["Frame", "Time_ms", "ROI01", "ROI02", "ROI03", "ROI04", "ROI05", "ROI06"])
            self.assertEqual(signal_table.signals.shape, (3, 6))
            self.assertEqual(signal_table.time_ms.tolist(), [0.0, 2.0, 4.0])
            np.testing.assert_allclose(signal_table.signals[0, :], [1, 2, 3, 4, 5, 6])
            np.testing.assert_allclose(signal_table.signals[2, :], [201, 202, 203, 204, 205, 206])

            export_dir = root / "exports"
            result = export_signal_documents(tiff_path, output_dir=export_dir, write_csv=True)
            self.assertTrue(result.xlsx_path.exists())
            self.assertTrue(result.csv_path and result.csv_path.exists())

            workbook = load_workbook(result.xlsx_path, data_only=True)
            try:
                self.assertEqual(workbook.sheetnames, ["Signals", "Metadata", "ROI_Map"])
                signals = workbook["Signals"]
                headers = [cell.value for cell in signals[1]]
                self.assertEqual(headers, signal_table.headers)
                self.assertEqual([signals.cell(2, col).value for col in range(1, 9)], [0, 0, 1, 2, 3, 4, 5, 6])
                self.assertEqual([signals.cell(4, col).value for col in range(1, 9)], [2, 4, 201, 202, 203, 204, 205, 206])

                metadata_rows = {
                    workbook["Metadata"].cell(row, 1).value: workbook["Metadata"].cell(row, 2).value
                    for row in range(2, workbook["Metadata"].max_row + 1)
                }
                self.assertEqual(metadata_rows["Sampling rate Hz"], 500)
                self.assertEqual(metadata_rows["Signal calculation"], "Mean raw pixel intensity inside each rectangular ROI")

                roi_map = workbook["ROI_Map"]
                self.assertEqual([roi_map.cell(2, col).value for col in range(1, 11)], ["ROI01", "rectangle", 1, 1, 0, 0, 10, 10, 10, 10])
                self.assertEqual([roi_map.cell(5, col).value for col in range(1, 11)], ["ROI04", "rectangle", 2, 1, 0, 10, 10, 20, 10, 10])
            finally:
                workbook.close()

            with result.csv_path.open(newline="", encoding="utf-8") as handle:
                rows = list(csv.reader(handle))
            self.assertEqual(rows[0], signal_table.headers)
            self.assertEqual(rows[2][:4], ["1", "2.0", "101.0", "102.0"])

    def test_extract_sampling_rate_supports_khz_metadata_comment(self) -> None:
        self.assertEqual(extract_sampling_rate_hz("1roi-2.08khz-plateau"), 2080.0)

    def test_export_signals_cli_exports_current_tiff(self) -> None:
        with tempfile.TemporaryDirectory() as temp_dir:
            root = Path(temp_dir)
            tiff_path = root / "sample.tiff"
            export_dir = root / "native_exports"
            metadata_path = metadata_path_for_tiff(tiff_path)
            self._write_synthetic_tiff(tiff_path)
            metadata_path.write_text(self._metadata_text("Comment: 6rois-500hz"), encoding="utf-8")

            exit_code = export_signals.main([str(tiff_path), "--output", str(export_dir)])

            self.assertEqual(exit_code, 0)
            self.assertTrue((export_dir / "sample_signals.xlsx").exists())
            self.assertTrue((export_dir / "sample_signals.csv").exists())

    def test_native_pipeline_writes_manifest_and_outputs(self) -> None:
        with tempfile.TemporaryDirectory() as temp_dir:
            root = Path(temp_dir)
            tiff_path = root / "sample.tiff"
            output_root = root / "native_pipeline"
            manifest_path = output_root / "manifest.json"
            metadata_path = metadata_path_for_tiff(tiff_path)
            self._write_synthetic_tiff(tiff_path)
            metadata_path.write_text(self._metadata_text("Comment: 6rois-500hz"), encoding="utf-8")

            exit_code = native_pipeline.main(
                [
                    str(tiff_path),
                    "--output",
                    str(output_root),
                    "--manifest",
                    str(manifest_path),
                ]
            )

            self.assertEqual(exit_code, 0)
            self.assertTrue(manifest_path.exists())
            manifest = json.loads(manifest_path.read_text(encoding="utf-8"))
            self.assertTrue(manifest["ok"])
            self.assertEqual(manifest["summary"]["tiff_count"], 1)
            record = manifest["tiffs"][0]
            self.assertTrue(record["actions"]["average_png"]["ok"])
            self.assertTrue(Path(record["actions"]["average_png"]["path"]).exists())
            self.assertTrue(record["actions"]["signals"]["ok"])
            self.assertTrue(Path(record["actions"]["signals"]["xlsx"]).exists())
            self.assertTrue(Path(record["actions"]["signals"]["csv"]).exists())
            self.assertTrue(record["actions"]["split"]["ok"])
            self.assertTrue(Path(record["actions"]["split"]["output_dir"]).exists())
            self.assertTrue(record["actions"]["motion_correction"]["skipped"])
            self.assertTrue(record["actions"]["roi_detection"]["skipped"])

    def test_native_pipeline_accepts_multiple_input_paths_for_queue_runs(self) -> None:
        with tempfile.TemporaryDirectory() as temp_dir:
            root = Path(temp_dir)
            first_tiff = root / "first.tiff"
            second_tiff = root / "second.tiff"
            output_root = root / "native_pipeline"
            manifest_path = output_root / "manifest.json"
            self._write_synthetic_tiff(first_tiff)
            self._write_synthetic_tiff(second_tiff)

            exit_code = native_pipeline.main(
                [
                    str(first_tiff),
                    str(second_tiff),
                    str(first_tiff),
                    "--output",
                    str(output_root),
                    "--manifest",
                    str(manifest_path),
                    "--no-average",
                    "--no-signals",
                    "--no-split",
                ]
            )

            self.assertEqual(exit_code, 0)
            manifest = json.loads(manifest_path.read_text(encoding="utf-8"))
            self.assertTrue(manifest["ok"])
            self.assertEqual(manifest["summary"]["tiff_count"], 2)
            self.assertEqual([Path(record["tiff"]).name for record in manifest["tiffs"]], ["first.tiff", "second.tiff"])
            self.assertEqual([Path(path).name for path in manifest["inputs"]], ["first.tiff", "second.tiff", "first.tiff"])

    def test_native_pipeline_runs_optional_hooks(self) -> None:
        with tempfile.TemporaryDirectory() as temp_dir:
            root = Path(temp_dir)
            hook_dir = root / "hooks"
            hook_dir.mkdir()
            hook_file = hook_dir / "pipeline_test_hooks.py"
            hook_file.write_text(
                """
from pathlib import Path

def motion(tiff_path, output_dir, stage, record):
    output_dir = Path(output_dir)
    output_dir.mkdir(parents=True, exist_ok=True)
    marker = output_dir / "motion_marker.txt"
    marker.write_text(f"{stage}:{Path(tiff_path).name}", encoding="utf-8")
    return {"ok": True, "marker": str(marker)}

def roi(tiff_path, output_dir, **kwargs):
    return {"ok": True, "detected_roi_count": 7, "source": str(tiff_path)}
""".lstrip(),
                encoding="utf-8",
            )
            sys.path.insert(0, str(hook_dir))
            try:
                tiff_path = root / "sample.tiff"
                output_root = root / "native_pipeline"
                manifest_path = output_root / "manifest.json"
                self._write_synthetic_tiff(tiff_path)

                exit_code = native_pipeline.main(
                    [
                        str(tiff_path),
                        "--output",
                        str(output_root),
                        "--manifest",
                        str(manifest_path),
                        "--no-average",
                        "--no-signals",
                        "--no-split",
                        "--motion-hook",
                        "pipeline_test_hooks:motion",
                        "--roi-hook",
                        "pipeline_test_hooks:roi",
                    ]
                )
            finally:
                sys.path.remove(str(hook_dir))

            self.assertEqual(exit_code, 0)
            manifest = json.loads(manifest_path.read_text(encoding="utf-8"))
            record = manifest["tiffs"][0]
            self.assertTrue(record["actions"]["motion_correction"]["ok"])
            self.assertFalse(record["actions"]["motion_correction"]["skipped"])
            self.assertTrue(Path(record["actions"]["motion_correction"]["marker"]).exists())
            self.assertTrue(record["actions"]["roi_detection"]["ok"])
            self.assertEqual(record["actions"]["roi_detection"]["detected_roi_count"], 7)

    def test_native_pipeline_runs_hook_from_python_file_path(self) -> None:
        with tempfile.TemporaryDirectory() as temp_dir:
            root = Path(temp_dir)
            hook_file = root / "external_motion_hook.py"
            hook_file.write_text(
                """
from pathlib import Path

def motion(tiff_path, output_dir, **kwargs):
    output_dir = Path(output_dir)
    output_dir.mkdir(parents=True, exist_ok=True)
    marker = output_dir / "external_motion.txt"
    marker.write_text(Path(tiff_path).name, encoding="utf-8")
    return {"ok": True, "marker": str(marker), "kind": "file-hook"}
""".lstrip(),
                encoding="utf-8",
            )

            tiff_path = root / "sample.tiff"
            output_root = root / "native_pipeline"
            manifest_path = output_root / "manifest.json"
            self._write_synthetic_tiff(tiff_path)

            exit_code = native_pipeline.main(
                [
                    str(tiff_path),
                    "--output",
                    str(output_root),
                    "--manifest",
                    str(manifest_path),
                    "--no-average",
                    "--no-signals",
                    "--no-split",
                    "--motion-hook",
                    f"{hook_file}:motion",
                ]
            )

            self.assertEqual(exit_code, 0)
            manifest = json.loads(manifest_path.read_text(encoding="utf-8"))
            motion = manifest["tiffs"][0]["actions"]["motion_correction"]
            self.assertTrue(motion["ok"])
            self.assertFalse(motion["skipped"])
            self.assertEqual(motion["kind"], "file-hook")
            self.assertTrue(Path(motion["marker"]).exists())

    def test_motion_hook_returned_tiff_is_used_downstream(self) -> None:
        with tempfile.TemporaryDirectory() as temp_dir:
            root = Path(temp_dir)
            hook_file = root / "motion_outputs_tiff.py"
            hook_file.write_text(
                """
from pathlib import Path
import numpy as np
from PIL import Image

def motion(tiff_path, output_dir, **kwargs):
    output_dir = Path(output_dir)
    output_dir.mkdir(parents=True, exist_ok=True)
    corrected = output_dir / "corrected.tiff"
    frames = []
    with Image.open(tiff_path) as image:
        for frame_index in range(getattr(image, "n_frames", 1)):
            image.seek(frame_index)
            frames.append(Image.fromarray(np.asarray(image, dtype=np.uint16) + 1000))
    frames[0].save(corrected, save_all=True, append_images=frames[1:])
    return {"ok": True, "corrected_tiff": str(corrected)}
""".lstrip(),
                encoding="utf-8",
            )

            tiff_path = root / "sample.tiff"
            metadata_path = metadata_path_for_tiff(tiff_path)
            output_root = root / "native_pipeline"
            manifest_path = output_root / "manifest.json"
            self._write_synthetic_tiff(tiff_path)
            metadata_path.write_text(self._metadata_text("Comment: 6rois-500hz"), encoding="utf-8")

            exit_code = native_pipeline.main(
                [
                    str(tiff_path),
                    "--output",
                    str(output_root),
                    "--manifest",
                    str(manifest_path),
                    "--no-average",
                    "--no-split",
                    "--motion-hook",
                    f"{hook_file}:motion",
                ]
            )

            self.assertEqual(exit_code, 0)
            manifest = json.loads(manifest_path.read_text(encoding="utf-8"))
            record = manifest["tiffs"][0]
            corrected_tiff = output_root / "sample" / "corrected.tiff"
            self.assertEqual(Path(record["working_tiff"]), corrected_tiff)
            self.assertTrue(metadata_path_for_tiff(corrected_tiff).exists())
            signals = record["actions"]["signals"]
            self.assertEqual(Path(signals["source_tiff"]), corrected_tiff)
            workbook = load_workbook(Path(signals["xlsx"]), data_only=True)
            try:
                self.assertEqual(workbook["Signals"].cell(2, 3).value, 1001)
            finally:
                workbook.close()

    def test_native_pipeline_invalid_hook_fails_before_outputs(self) -> None:
        with tempfile.TemporaryDirectory() as temp_dir:
            root = Path(temp_dir)
            hook_file = root / "bad_hook.py"
            hook_file.write_text(
                """
def existing(**kwargs):
    return {"ok": True}
""".lstrip(),
                encoding="utf-8",
            )
            tiff_path = root / "sample.tiff"
            output_root = root / "native_pipeline"
            manifest_path = output_root / "manifest.json"
            self._write_synthetic_tiff(tiff_path)

            exit_code = native_pipeline.main(
                [
                    str(tiff_path),
                    "--output",
                    str(output_root),
                    "--manifest",
                    str(manifest_path),
                    "--motion-hook",
                    f"{hook_file}:missing",
                ]
            )

            self.assertEqual(exit_code, 2)
            self.assertFalse(manifest_path.exists())
            self.assertFalse(output_root.exists())

    def test_native_pipeline_runtime_hook_failure_skips_downstream_outputs(self) -> None:
        with tempfile.TemporaryDirectory() as temp_dir:
            root = Path(temp_dir)
            hook_file = root / "failing_hook.py"
            hook_file.write_text(
                """
def motion(**kwargs):
    return {"ok": False, "reason": "synthetic failure"}
""".lstrip(),
                encoding="utf-8",
            )
            tiff_path = root / "sample.tiff"
            output_root = root / "native_pipeline"
            manifest_path = output_root / "manifest.json"
            self._write_synthetic_tiff(tiff_path)

            exit_code = native_pipeline.main(
                [
                    str(tiff_path),
                    "--output",
                    str(output_root),
                    "--manifest",
                    str(manifest_path),
                    "--motion-hook",
                    f"{hook_file}:motion",
                ]
            )

            self.assertEqual(exit_code, 1)
            manifest = json.loads(manifest_path.read_text(encoding="utf-8"))
            record = manifest["tiffs"][0]
            self.assertFalse(record["ok"])
            self.assertFalse(record["actions"]["motion_correction"]["ok"])
            self.assertNotIn("signals", record["actions"])
            self.assertNotIn("average_png", record["actions"])
            self.assertNotIn("split", record["actions"])

    def test_validate_hooks_accepts_file_hook_and_rejects_missing_function(self) -> None:
        with tempfile.TemporaryDirectory() as temp_dir:
            root = Path(temp_dir)
            hook_file = root / "validate_hook.py"
            hook_file.write_text(
                """
def motion(**kwargs):
    return {"ok": True}
""".lstrip(),
                encoding="utf-8",
            )

            self.assertEqual(validate_hooks.main(["--motion-hook", f"{hook_file}:motion"]), 0)
            self.assertEqual(validate_hooks.main(["--motion-hook", f"{hook_file}:missing"]), 1)

    @staticmethod
    def _write_synthetic_tiff(path: Path) -> None:
        frames = []
        for frame_number in range(3):
            frame = np.zeros((20, 30), dtype=np.uint16)
            for roi_index in range(1, 7):
                zero_based = roi_index - 1
                row = zero_based // 3
                column = zero_based % 3
                left = column * 10
                upper = row * 10
                frame[upper : upper + 10, left : left + 10] = frame_number * 100 + roi_index
            frames.append(Image.fromarray(frame))
        frames[0].save(path, save_all=True, append_images=frames[1:])

    @staticmethod
    def _metadata_text(comment_line: str) -> str:
        return f"""%Metadata for:
FileName: synthetic.mesc
MestagMemo: sample
Type: MultiROIChessBoard
{comment_line}

>Context start
%SIZE
 Width: 30
 Height: 20
 Average: 2
%PATTERN
 PatternNum: 2
 RoiCount: 6
 RoiSize (um): [[10, 10], [10, 10], [10, 10], [10, 10], [10, 10], [10, 10]]
 RoiIndex: [1, 2, 3, 4, 5, 6]
 Channel: UG
 DataType: uint16
<Context end
"""


if __name__ == "__main__":
    unittest.main()
