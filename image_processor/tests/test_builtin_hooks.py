import json
import os
import tempfile
import unittest
from pathlib import Path

import numpy as np
from openpyxl import load_workbook
from PIL import Image

from image_processor import native_pipeline, validate_hooks
from image_processor.motion_correction import auto_max_shift_px, estimate_phase_shift, motion_hook
from image_processor.roi_core import RoiBox, metadata_path_for_tiff
from image_processor.segmentation_hooks import mask_boundary, roi_hook as segmentation_roi_hook
from image_processor.tiff_backend import iter_tiff_frames
from image_processor.tracked_segmentation import (
    TrackMask,
    _track_masks_and_write_debug,
    estimate_mask_shift,
    roi_hook as tracked_roi_hook,
)


class BuiltInHookTests(unittest.TestCase):
    def test_motion_hook_writes_corrected_tiff_and_qc(self) -> None:
        with tempfile.TemporaryDirectory() as temp_dir:
            root = Path(temp_dir)
            tiff_path = root / "motion_sample.tiff"
            metadata_path = metadata_path_for_tiff(tiff_path)
            self._write_motion_tiff(tiff_path)
            metadata_path.write_text(self._metadata_text(width=32, height=32, roi_count=1), encoding="utf-8")

            result = motion_hook(tiff_path, output_dir=root / "out", max_shift_px=5)

            self.assertTrue(result["ok"])
            corrected_tiff = Path(result["corrected_tiff"])
            self.assertTrue(corrected_tiff.exists())
            self.assertTrue(Path(result["qc_json"]).exists())
            self.assertTrue(Path(result["shifts_csv"]).exists())
            self.assertTrue(metadata_path_for_tiff(corrected_tiff).exists())

            corrected_frames = [np.asarray(frame) for frame in iter_tiff_frames(corrected_tiff)]
            self.assertEqual(len(corrected_frames), 2)
            np.testing.assert_array_equal(corrected_frames[0], corrected_frames[1])

    def test_motion_hook_skips_single_frame_tiff_with_qc(self) -> None:
        with tempfile.TemporaryDirectory() as temp_dir:
            root = Path(temp_dir)
            tiff_path = root / "still_sample.tiff"
            metadata_path = metadata_path_for_tiff(tiff_path)
            self._write_single_frame_tiff(tiff_path)
            metadata_path.write_text(self._metadata_text(width=32, height=16, roi_count=2), encoding="utf-8")

            result = motion_hook(tiff_path, output_dir=root / "out", max_shift_px=5)

            self.assertTrue(result["ok"])
            self.assertTrue(result["motion_skipped"])
            self.assertEqual(result["frame_count"], 1)
            self.assertEqual(result["layout_source"], "metadata_tiles")
            self.assertNotIn("corrected_tiff", result)
            self.assertNotIn("shifts_csv", result)
            qc_path = Path(result["qc_json"])
            self.assertTrue(qc_path.exists())
            qc = json.loads(qc_path.read_text(encoding="utf-8"))
            self.assertTrue(qc["motion_skipped"])
            self.assertIn("Single-frame TIFF", qc["reason"])

    def test_motion_auto_shift_scales_down_for_tiny_metadata_tiles(self) -> None:
        with tempfile.TemporaryDirectory() as temp_dir:
            root = Path(temp_dir)
            tiff_path = root / "tiny_tiles.tiff"
            metadata_path = metadata_path_for_tiff(tiff_path)
            first = Image.fromarray(np.full((10, 20), 10, dtype=np.uint16))
            second = Image.fromarray(np.full((10, 20), 11, dtype=np.uint16))
            first.save(tiff_path, save_all=True, append_images=[second])
            metadata_path.write_text(self._metadata_text(width=20, height=10, roi_count=2), encoding="utf-8")

            result = motion_hook(tiff_path, output_dir=root / "out")

            self.assertTrue(result["ok"])
            self.assertEqual(result["max_shift_px"], 1)

    def test_motion_auto_shift_allows_larger_tiles_to_move_more(self) -> None:
        from image_processor.roi_core import RoiBox

        boxes = [
            RoiBox(
                ordinal=1,
                roi_index=1,
                row=0,
                column=0,
                left=0,
                upper=0,
                right=64,
                lower=64,
            )
        ]

        self.assertEqual(auto_max_shift_px(boxes), 5)

    def test_shared_template_phase_shift_estimates_known_shared_motion(self) -> None:
        reference = np.zeros((20, 30), dtype=np.float32)
        reference[5:10, 12:18] = 100
        moving = np.zeros_like(reference)
        moving[6:11, 10:16] = 100

        shift = estimate_phase_shift(reference, moving, max_shift_px=3, subpixel=False, min_peak_ratio=1.0)

        self.assertTrue(shift["accepted"])
        self.assertEqual(shift["dy"], -1.0)
        self.assertEqual(shift["dx"], 2.0)

    def test_motion_hook_shared_template_writes_extended_qc(self) -> None:
        with tempfile.TemporaryDirectory() as temp_dir:
            root = Path(temp_dir)
            tiff_path = root / "motion_sample.tiff"
            metadata_path = metadata_path_for_tiff(tiff_path)
            self._write_motion_tiff(tiff_path)
            metadata_path.write_text(self._metadata_text(width=32, height=32, roi_count=1), encoding="utf-8")

            result = motion_hook(
                tiff_path,
                output_dir=root / "out",
                method="shared_template",
                max_shift_px=5,
                subpixel=True,
                min_peak_ratio=1.0,
            )

            self.assertTrue(result["ok"])
            self.assertEqual(result["method"], "shared_template")
            self.assertTrue(Path(result["corrected_tiff"]).exists())
            shifts_csv = Path(result["shifts_csv"])
            self.assertTrue(shifts_csv.exists())
            header = shifts_csv.read_text(encoding="utf-8").splitlines()[0]
            self.assertIn("peak_ratio", header)
            self.assertIn("estimated_dy", header)

    def test_motion_hook_shared_template_residual_writes_staged_qc(self) -> None:
        with tempfile.TemporaryDirectory() as temp_dir:
            root = Path(temp_dir)
            tiff_path = root / "motion_sample.tiff"
            metadata_path = metadata_path_for_tiff(tiff_path)
            self._write_motion_tiff(tiff_path)
            metadata_path.write_text(self._metadata_text(width=32, height=32, roi_count=1), encoding="utf-8")

            result = motion_hook(
                tiff_path,
                output_dir=root / "out",
                method="shared_template_residual",
                max_shift_px=5,
                subpixel=False,
                residual_max_shift_px=1,
                residual_min_peak_ratio=1.1,
            )

            self.assertTrue(result["ok"])
            self.assertEqual(result["method"], "shared_template_residual")
            self.assertTrue(Path(result["corrected_tiff"]).exists())
            self.assertIn("residual_moved_shift_count", result)
            shifts_csv = Path(result["shifts_csv"])
            header = shifts_csv.read_text(encoding="utf-8").splitlines()[0]
            self.assertIn("stage", header)
            self.assertIn("residual_min_peak_ratio", header)

    def test_roi_hook_segments_inside_each_metadata_tile(self) -> None:
        with tempfile.TemporaryDirectory() as temp_dir:
            root = Path(temp_dir)
            tiff_path = root / "roi_sample.tiff"
            metadata_path = metadata_path_for_tiff(tiff_path)
            self._write_roi_tiff(tiff_path)
            metadata_path.write_text(self._metadata_text(width=32, height=16, roi_count=2), encoding="utf-8")

            result = segmentation_roi_hook(tiff_path, output_dir=root / "out", min_area=4)

            self.assertTrue(result["ok"])
            self.assertEqual(result["detected_roi_count"], 2)
            self.assertEqual(Path(result["metadata_path"]), metadata_path)
            self.assertTrue(Path(result["mask_path"]).exists())
            self.assertTrue(Path(result["label_tiff"]).exists())
            self.assertTrue(Path(result["qc_overlay_png"]).exists())
            self.assertTrue(Path(result["summary_mean_png"]).exists())
            self.assertTrue(Path(result["summary_std_png"]).exists())
            self.assertTrue(Path(result["summary_max_png"]).exists())
            self.assertTrue(Path(result["summary_manifest_path"]).exists())
            with Image.open(result["qc_overlay_png"]) as overlay:
                self.assertGreater(overlay.size[0], 32)
                self.assertGreater(overlay.size[1], 16)
                self.assertNotEqual(overlay.getpixel((0, 0)), (0, 210, 255))

            with np.load(result["mask_path"], allow_pickle=False) as masks_file:
                self.assertEqual(masks_file["masks"].shape, (2, 16, 32))
                self.assertEqual(int(np.max(masks_file["label_image"])), 2)
                self.assertEqual(masks_file["labels"].tolist(), ["ROI01_S01", "ROI02_S01"])

            manifest = json.loads(Path(result["mask_manifest_path"]).read_text(encoding="utf-8"))
            self.assertEqual(manifest["layout_source"], "metadata_tiles")
            self.assertEqual(manifest["detected_roi_count"], 2)
            self.assertTrue(Path(manifest["qc_overlay_png"]).exists())
            self.assertEqual(
                [(tile["roi_index"], tile["detected_roi_count"]) for tile in manifest["tiles"]],
                [(1, 1), (2, 1)],
            )
            self.assertEqual(
                set(manifest["summary_outputs"]),
                {"summary_mean_png", "summary_std_png", "summary_max_png", "summary_manifest_path"},
            )
            self.assertEqual(
                {entry["parent_roi_index"] for entry in manifest["detections"]},
                {1, 2},
            )

    def test_roi_hook_segments_broad_tile_foreground_not_only_hot_pixels(self) -> None:
        with tempfile.TemporaryDirectory() as temp_dir:
            root = Path(temp_dir)
            tiff_path = root / "small_tile_foreground.tiff"
            metadata_path = metadata_path_for_tiff(tiff_path)

            frame = np.full((10, 20), 10, dtype=np.uint16)
            frame[2:8, 1:5] = 30
            frame[3, 2] = 90
            frame[1:4, 12:19] = 35
            frame[2, 16] = 120
            Image.fromarray(frame).save(tiff_path)
            metadata_path.write_text(self._metadata_text(width=20, height=10, roi_count=2), encoding="utf-8")

            result = segmentation_roi_hook(tiff_path, output_dir=root / "out")

            self.assertTrue(result["ok"])
            self.assertEqual(result["detected_roi_count"], 2)
            manifest = json.loads(Path(result["mask_manifest_path"]).read_text(encoding="utf-8"))
            self.assertEqual(manifest["parameters"]["threshold_method"], "foreground")
            areas = [int(entry["area_px"]) for entry in manifest["detections"]]
            self.assertGreaterEqual(min(areas), 20)
            self.assertEqual(
                [(tile["roi_index"], tile["detected_roi_count"]) for tile in manifest["tiles"]],
                [(1, 1), (2, 1)],
            )

    def test_mask_boundary_marks_contour_not_interior(self) -> None:
        mask = np.zeros((5, 5), dtype=bool)
        mask[1:4, 1:4] = True

        boundary = mask_boundary(mask)

        self.assertTrue(boundary[1, 1])
        self.assertTrue(boundary[3, 3])
        self.assertFalse(boundary[2, 2])
        self.assertEqual(int(np.sum(boundary)), 8)

    def test_builtin_hook_specs_validate(self) -> None:
        self.assertEqual(
            validate_hooks.main(
                [
                    "--motion-hook",
                    "image_processor.motion_correction:motion_hook",
                    "--motion-hook-param",
                    "max_shift_px=5",
                    "--roi-hook",
                    "image_processor.segmentation_hooks:roi_hook",
                    "--roi-hook-param",
                    "min_area=4",
                ]
            ),
            0,
        )

    def test_native_pipeline_uses_roi_hook_masks_for_signal_export(self) -> None:
        with tempfile.TemporaryDirectory() as temp_dir:
            root = Path(temp_dir)
            tiff_path = root / "roi_sample.tiff"
            metadata_path = metadata_path_for_tiff(tiff_path)
            output_root = root / "native_pipeline"
            manifest_path = output_root / "manifest.json"
            self._write_roi_tiff(tiff_path)
            metadata_path.write_text(self._metadata_text(width=32, height=16, roi_count=2), encoding="utf-8")

            exit_code = native_pipeline.main(
                [
                    str(tiff_path),
                    "--output",
                    str(output_root),
                    "--manifest",
                    str(manifest_path),
                    "--sampling-rate",
                    "100",
                    "--no-average",
                    "--no-split",
                    "--roi-hook",
                    "image_processor.segmentation_hooks:roi_hook",
                ]
            )

            self.assertEqual(exit_code, 0)
            manifest = json.loads(manifest_path.read_text(encoding="utf-8"))
            record = manifest["tiffs"][0]
            self.assertEqual(record["actions"]["signals"]["signal_source"], "mask")
            self.assertEqual(record["actions"]["signals"]["roi_count"], 2)
            self.assertTrue(Path(record["actions"]["signals"]["mask_path"]).exists())
            self.assertTrue(Path(record["actions"]["signals"]["mask_manifest_path"]).exists())
            roi_action = record["actions"]["roi_detection"]
            self.assertTrue(Path(roi_action["qc_overlay_png"]).exists())
            self.assertTrue(Path(roi_action["summary_mean_png"]).exists())
            self.assertTrue(Path(roi_action["summary_std_png"]).exists())
            self.assertTrue(Path(roi_action["summary_max_png"]).exists())
            self.assertTrue(Path(roi_action["summary_manifest_path"]).exists())

            workbook = load_workbook(Path(record["actions"]["signals"]["xlsx"]), data_only=True)
            try:
                self.assertEqual(workbook.sheetnames, ["Signals", "Metadata", "Mask_Map"])
                signals = workbook["Signals"]
                headers = [cell.value for cell in signals[1]]
                self.assertEqual(headers, ["Frame", "Time_ms", "ROI01_S01", "ROI02_S01"])
                self.assertEqual([signals.cell(2, col).value for col in range(1, 5)], [0, 0, 100, 120])
                self.assertEqual([signals.cell(4, col).value for col in range(1, 5)], [2, 20, 102, 122])

                metadata_rows = {
                    workbook["Metadata"].cell(row, 1).value: workbook["Metadata"].cell(row, 2).value
                    for row in range(2, workbook["Metadata"].max_row + 1)
                }
                self.assertEqual(
                    metadata_rows["Signal calculation"],
                    "Mean raw pixel intensity inside segmentation masks",
                )
                self.assertEqual(metadata_rows["Mask count"], 2)
                self.assertEqual(workbook["Mask_Map"].cell(2, 1).value, "ROI01_S01")
                self.assertEqual(workbook["Mask_Map"].cell(3, 1).value, "ROI02_S01")
            finally:
                workbook.close()

    def test_native_pipeline_can_background_correct_mask_signals(self) -> None:
        with tempfile.TemporaryDirectory() as temp_dir:
            root = Path(temp_dir)
            tiff_path = root / "roi_sample.tiff"
            metadata_path = metadata_path_for_tiff(tiff_path)
            output_root = root / "native_pipeline"
            manifest_path = output_root / "manifest.json"
            self._write_roi_tiff(tiff_path)
            metadata_path.write_text(self._metadata_text(width=32, height=16, roi_count=2), encoding="utf-8")

            exit_code = native_pipeline.main(
                [
                    str(tiff_path),
                    "--output",
                    str(output_root),
                    "--manifest",
                    str(manifest_path),
                    "--sampling-rate",
                    "100",
                    "--mask-background",
                    "parent_tile",
                    "--no-average",
                    "--no-split",
                    "--roi-hook",
                    "image_processor.segmentation_hooks:roi_hook",
                ]
            )

            self.assertEqual(exit_code, 0)
            manifest = json.loads(manifest_path.read_text(encoding="utf-8"))
            record = manifest["tiffs"][0]
            self.assertEqual(manifest["settings"]["mask_background"], "parent_tile")
            self.assertEqual(record["actions"]["signals"]["background_mode"], "parent_tile")

            workbook = load_workbook(Path(record["actions"]["signals"]["xlsx"]), data_only=True)
            try:
                self.assertEqual(workbook.sheetnames, ["Signals", "Background", "Metadata", "Mask_Map"])
                signals = workbook["Signals"]
                self.assertEqual([signals.cell(2, col).value for col in range(1, 5)], [0, 0, 90, 110])
                self.assertEqual([signals.cell(4, col).value for col in range(1, 5)], [2, 20, 90, 110])
                background = workbook["Background"]
                self.assertEqual([background.cell(2, col).value for col in range(1, 5)], [0, 0, 10, 10])
                self.assertEqual([background.cell(4, col).value for col in range(1, 5)], [2, 20, 12, 12])
                metadata_rows = {
                    workbook["Metadata"].cell(row, 1).value: workbook["Metadata"].cell(row, 2).value
                    for row in range(2, workbook["Metadata"].max_row + 1)
                }
                self.assertEqual(metadata_rows["Background mode"], "parent_tile")
            finally:
                workbook.close()

    def test_native_pipeline_exports_signal_comparison_csvs_for_motion_and_segmentation(self) -> None:
        with tempfile.TemporaryDirectory() as temp_dir:
            root = Path(temp_dir)
            tiff_path = root / "roi_sample.tiff"
            metadata_path = metadata_path_for_tiff(tiff_path)
            output_root = root / "native_pipeline"
            manifest_path = output_root / "manifest.json"
            self._write_roi_tiff(tiff_path)
            metadata_path.write_text(self._metadata_text(width=32, height=16, roi_count=2), encoding="utf-8")

            exit_code = native_pipeline.main(
                [
                    str(tiff_path),
                    "--output",
                    str(output_root),
                    "--manifest",
                    str(manifest_path),
                    "--sampling-rate",
                    "100",
                    "--no-average",
                    "--no-split",
                    "--motion-hook",
                    "image_processor.motion_correction:motion_hook",
                    "--motion-hook-param",
                    "method=shared_template_residual",
                    "--motion-hook-param",
                    "max_shift_px=2",
                    "--motion-hook-param",
                    "residual_max_shift_px=1",
                    "--motion-hook-param",
                    "residual_min_peak_ratio=1.10",
                    "--roi-hook",
                    "image_processor.segmentation_hooks:roi_hook",
                ]
            )

            self.assertEqual(exit_code, 0)
            manifest = json.loads(manifest_path.read_text(encoding="utf-8"))
            record = manifest["tiffs"][0]
            self.assertTrue(Path(record["actions"]["roi_detection_raw"]["mask_path"]).exists())

            comparison = record["actions"]["signal_comparison_csvs"]
            self.assertTrue(comparison["ok"])
            self.assertTrue(Path(comparison["xlsx"]).exists())
            self.assertEqual(
                set(comparison["variants"]),
                {
                    "raw",
                    "tracked_segmentation",
                    "motion_corrected",
                    "segmentation",
                    "motion_segmentation_on_raw",
                    "motion_corrected_segmentation",
                },
            )
            self.assertTrue(comparison["variants"]["tracked_segmentation"]["skipped"])
            for variant_name in [
                "raw",
                "motion_corrected",
                "segmentation",
                "motion_segmentation_on_raw",
                "motion_corrected_segmentation",
            ]:
                variant = comparison["variants"][variant_name]
                self.assertFalse(variant.get("skipped", False), variant.get("reason"))
                self.assertTrue(Path(variant["csv"]).exists())
            raw_tiff_path = Path(comparison["variants"]["motion_segmentation_on_raw"]["source_tiff"])
            mask_source_path = Path(comparison["variants"]["motion_segmentation_on_raw"]["mask_source_tiff"])
            self.assertEqual(raw_tiff_path, tiff_path)
            self.assertNotEqual(mask_source_path, tiff_path)
            self.assertEqual(len(comparison["csvs"]), 5)

            comparison_workbook = load_workbook(Path(comparison["xlsx"]), data_only=True)
            try:
                self.assertEqual(
                    comparison_workbook.sheetnames,
                    [
                        "Raw",
                        "Tracked Segmentation",
                        "Segmented",
                        "Motion Seg on Raw",
                        "Motion Corrected",
                        "Segmented+Motion",
                        "Export Map",
                    ],
                )
                self.assertEqual(comparison_workbook["Raw"].cell(1, 1).value, "Frame")
                self.assertEqual(comparison_workbook["Tracked Segmentation"].cell(2, 1).value, "Skipped")
                self.assertEqual(comparison_workbook["Segmented"].cell(1, 3).value, "ROI01_S01")
                self.assertEqual(comparison_workbook["Motion Seg on Raw"].cell(1, 3).value, "ROI01_S01")
                self.assertEqual(comparison_workbook["Motion Corrected"].cell(1, 1).value, "Frame")
                self.assertEqual(comparison_workbook["Segmented+Motion"].cell(1, 3).value, "ROI01_S01")
            finally:
                comparison_workbook.close()

            qc_images = record["actions"]["qc_images"]
            self.assertTrue(qc_images["ok"])
            self.assertTrue(Path(qc_images["raw_average_png"]).exists())
            self.assertTrue(Path(qc_images["motion_vs_raw_qc_png"]).exists())
            self.assertTrue(Path(qc_images["segmentation_overlay_png"]).exists())
            self.assertTrue(Path(qc_images["raw_segmentation_overlay_png"]).exists())
            with Image.open(qc_images["raw_average_png"]) as qc_image:
                self.assertGreater(qc_image.size[0], 32)
                self.assertGreater(qc_image.size[1], 16)

    def test_tracked_shift_estimator_moves_mask_and_clips_to_parent_roi(self) -> None:
        box = RoiBox(ordinal=1, roi_index=1, row=0, column=0, left=0, upper=0, right=8, lower=8)
        mask = TrackMask(
            label="ROI01_S01",
            label_id=1,
            parent_roi_index=1,
            box=box,
            ys=np.asarray([2, 2, 3, 3], dtype=np.int32),
            xs=np.asarray([2, 3, 2, 3], dtype=np.int32),
            template_values=np.full(4, 100.0, dtype=np.float64),
        )
        frame = np.zeros((8, 8), dtype=np.float32)
        frame[3:5, 3:5] = 100.0

        track = estimate_mask_shift(frame, mask, max_shift_px=2)

        self.assertEqual((track.dx, track.dy), (1, 1))
        self.assertTrue(track.valid_frame)
        self.assertEqual(track.visible_pixel_count, 4)
        self.assertEqual(track.visible_fraction, 1.0)
        self.assertEqual(track.signal, 100.0)

        edge_mask = TrackMask(
            label="ROI01_S02",
            label_id=2,
            parent_roi_index=1,
            box=box,
            ys=np.asarray([2, 2, 3, 3], dtype=np.int32),
            xs=np.asarray([6, 7, 6, 7], dtype=np.int32),
            template_values=np.full(4, 100.0, dtype=np.float64),
        )
        edge_frame = np.zeros((8, 8), dtype=np.float32)
        edge_frame[2:4, 7] = 100.0

        edge_track = estimate_mask_shift(edge_frame, edge_mask, max_shift_px=1)

        self.assertEqual((edge_track.dx, edge_track.dy), (1, 0))
        self.assertTrue(edge_track.valid_frame)
        self.assertTrue(edge_track.trace_valid)
        self.assertTrue(edge_track.clipped)
        self.assertEqual(edge_track.visible_pixel_count, 2)
        self.assertEqual(edge_track.visible_fraction, 0.5)
        self.assertTrue(edge_track.edge_touch)
        self.assertEqual(edge_track.signal, 100.0)

        excluded_edge_track = estimate_mask_shift(
            edge_frame,
            edge_mask,
            max_shift_px=1,
            exclude_clipped_masks=True,
        )

        self.assertEqual((excluded_edge_track.dx, excluded_edge_track.dy), (1, 0))
        self.assertTrue(excluded_edge_track.valid_frame)
        self.assertFalse(excluded_edge_track.trace_valid)
        self.assertEqual(excluded_edge_track.trace_excluded_reason, "clipped_mask")
        self.assertTrue(np.isnan(excluded_edge_track.signal))

    def test_tracked_signal_export_uses_visible_pixels_when_mask_is_clipped(self) -> None:
        with tempfile.TemporaryDirectory() as temp_dir:
            root = Path(temp_dir)
            tiff_path = root / "clipped_sample.tiff"
            frames = []
            for offset in (0, 1):
                frame = np.zeros((8, 8), dtype=np.uint16)
                frame[2:4, 6 + offset : 8 + offset] = 100
                frames.append(Image.fromarray(frame))
            frames[0].save(tiff_path, save_all=True, append_images=frames[1:])

            box = RoiBox(ordinal=1, roi_index=1, row=0, column=0, left=0, upper=0, right=8, lower=8)
            mask = TrackMask(
                label="ROI01_S01",
                label_id=1,
                parent_roi_index=1,
                box=box,
                ys=np.asarray([2, 2, 3, 3], dtype=np.int32),
                xs=np.asarray([6, 7, 6, 7], dtype=np.int32),
                template_values=np.full(4, 100.0, dtype=np.float64),
            )

            result = _track_masks_and_write_debug(
                tiff_path=tiff_path,
                output_dir=root / "tracked",
                boxes=[box],
                track_masks=[mask],
                frame_count=2,
                sampling_rate_hz=100.0,
                max_shift_px=1,
                min_tracking_score=None,
                tracking_mode="fixed",
                min_visible_fraction=0.0,
                debug_tiles=False,
            )

            self.assertEqual(result["signals"][0, 0], 100.0)
            self.assertEqual(result["signals"][1, 0], 100.0)
            self.assertTrue(result["tracks"][1].valid_frame)
            self.assertTrue(result["tracks"][1].trace_valid)
            self.assertTrue(result["tracks"][1].clipped)
            self.assertEqual(result["tracks"][1].visible_pixel_count, 2)

            csv_lines = Path(result["trace_csv"]).read_text(encoding="utf-8").splitlines()
            self.assertTrue(csv_lines[1].endswith(",100.0"))
            self.assertTrue(csv_lines[2].endswith(",100.0"))

    def test_cumulative_tracking_follows_drift_beyond_fixed_search_window(self) -> None:
        with tempfile.TemporaryDirectory() as temp_dir:
            root = Path(temp_dir)
            tiff_path = root / "drift_sample.tiff"
            frames = []
            for frame_index in range(5):
                frame = np.zeros((12, 12), dtype=np.uint16)
                frame[4:6, 2 + frame_index : 4 + frame_index] = 100
                frames.append(Image.fromarray(frame))
            frames[0].save(tiff_path, save_all=True, append_images=frames[1:])

            box = RoiBox(ordinal=1, roi_index=1, row=0, column=0, left=0, upper=0, right=12, lower=12)
            mask = TrackMask(
                label="ROI01_S01",
                label_id=1,
                parent_roi_index=1,
                box=box,
                ys=np.asarray([4, 4, 5, 5], dtype=np.int32),
                xs=np.asarray([2, 3, 2, 3], dtype=np.int32),
                template_values=np.full(4, 100.0, dtype=np.float64),
            )

            fixed = _track_masks_and_write_debug(
                tiff_path=tiff_path,
                output_dir=root / "fixed",
                boxes=[box],
                track_masks=[mask],
                frame_count=5,
                sampling_rate_hz=100.0,
                max_shift_px=2,
                min_tracking_score=None,
                tracking_mode="fixed",
                min_visible_fraction=0.0,
                debug_tiles=False,
            )
            cumulative = _track_masks_and_write_debug(
                tiff_path=tiff_path,
                output_dir=root / "cumulative",
                boxes=[box],
                track_masks=[mask],
                frame_count=5,
                sampling_rate_hz=100.0,
                max_shift_px=2,
                min_tracking_score=None,
                tracking_mode="cumulative",
                local_shift_px=1,
                rescue_shift_px=2,
                min_visible_fraction=0.0,
                debug_tiles=False,
            )

            fixed_last = fixed["tracks"][-1]
            cumulative_last = cumulative["tracks"][-1]
            self.assertNotEqual((fixed_last.cumulative_dx, fixed_last.cumulative_dy), (4, 0))
            self.assertEqual((cumulative_last.cumulative_dx, cumulative_last.cumulative_dy), (4, 0))
            self.assertEqual(cumulative_last.track_state, "accepted")
            self.assertEqual(cumulative_last.signal, 100.0)

    def test_tracked_segmentation_hook_exports_tracks_signals_and_debug_tiff(self) -> None:
        with tempfile.TemporaryDirectory() as temp_dir:
            root = Path(temp_dir)
            tiff_path = root / "tracked_sample.tiff"
            metadata_path = metadata_path_for_tiff(tiff_path)
            self._write_tracked_tiff(tiff_path)
            metadata_path.write_text(self._metadata_text(width=16, height=16, roi_count=1), encoding="utf-8")

            result = tracked_roi_hook(
                tiff_path,
                output_dir=root / "tracked_out",
                min_area=4,
                max_shift_px=2,
                foreground_percentile=70,
                sampling_rate_hz=100,
            )

            self.assertTrue(result["ok"])
            self.assertEqual(result["mode"], "tracked_segmentation")
            self.assertEqual(result["initial_mask_source"], "motion_corrected_mean")
            self.assertEqual(result["requested_initial_mask_source"], "motion_corrected_mean")
            self.assertTrue(Path(result["initial_mask_tiff"]).exists())
            self.assertNotEqual(Path(result["initial_mask_tiff"]), tiff_path)
            self.assertTrue(Path(result["initial_mask_motion_qc_json"]).exists())
            self.assertGreaterEqual(result["detected_roi_count"], 1)
            self.assertTrue(Path(result["mask_path"]).exists())
            self.assertTrue(Path(result["tracked_manifest_path"]).exists())
            self.assertTrue(Path(result["tracked_tracks_csv"]).exists())
            self.assertTrue(Path(result["tracked_trace_csv"]).exists())
            self.assertTrue(Path(result["tracked_trace_xlsx"]).exists())
            self.assertEqual(len(result["tracked_debug_tiffs"]), 1)
            self.assertEqual(len(result["tracked_debug_raw_tiffs"]), 1)
            self.assertEqual(len(result["tracked_debug_mask_tiffs"]), 1)
            debug_tiff = Path(result["tracked_debug_tiffs"][0])
            self.assertTrue(debug_tiff.exists())
            with Image.open(debug_tiff) as image:
                self.assertEqual(image.size, (32, 16))
            with Image.open(result["tracked_debug_raw_tiffs"][0]) as image:
                self.assertEqual(image.size, (16, 16))
            with Image.open(result["tracked_debug_mask_tiffs"][0]) as image:
                self.assertEqual(image.size, (16, 16))

            tracks_header = Path(result["tracked_tracks_csv"]).read_text(encoding="utf-8").splitlines()[0]
            self.assertIn("visible_fraction", tracks_header)
            self.assertIn("edge_touch", tracks_header)

    def test_native_pipeline_dynamic_tracked_segmentation_exports_only_raw_and_tracked_comparisons(self) -> None:
        with tempfile.TemporaryDirectory() as temp_dir:
            root = Path(temp_dir)
            tiff_path = root / "tracked_sample.tiff"
            metadata_path = metadata_path_for_tiff(tiff_path)
            output_root = root / "native_pipeline"
            manifest_path = output_root / "manifest.json"
            self._write_tracked_tiff(tiff_path)
            metadata_path.write_text(self._metadata_text(width=16, height=16, roi_count=1), encoding="utf-8")

            exit_code = native_pipeline.main(
                [
                    str(tiff_path),
                    "--output",
                    str(output_root),
                    "--manifest",
                    str(manifest_path),
                    "--sampling-rate",
                    "100",
                    "--no-average",
                    "--no-split",
                    "--roi-hook",
                    "image_processor.tracked_segmentation:roi_hook",
                    "--roi-hook-param",
                    "min_area=4",
                    "--roi-hook-param",
                    "max_shift_px=2",
                    "--roi-hook-param",
                    "foreground_percentile=70",
                ]
            )

            self.assertEqual(exit_code, 0)
            manifest = json.loads(manifest_path.read_text(encoding="utf-8"))
            record = manifest["tiffs"][0]
            self.assertTrue(record["ok"])
            self.assertEqual(Path(record["working_tiff"]), tiff_path)
            self.assertTrue(record["actions"]["motion_correction"]["skipped"])

            roi_action = record["actions"]["roi_detection"]
            self.assertEqual(roi_action["mode"], "tracked_segmentation")
            self.assertEqual(roi_action["initial_mask_source"], "motion_corrected_mean")
            self.assertEqual(roi_action["requested_initial_mask_source"], "motion_corrected_mean")
            self.assertTrue(Path(roi_action["initial_mask_tiff"]).exists())
            self.assertNotEqual(Path(roi_action["initial_mask_tiff"]), tiff_path)
            self.assertTrue(Path(roi_action["initial_mask_motion_qc_json"]).exists())
            self.assertTrue(Path(roi_action["tracked_manifest_path"]).exists())
            self.assertTrue(Path(roi_action["tracked_tracks_csv"]).exists())
            self.assertTrue(Path(roi_action["tracked_trace_csv"]).exists())
            self.assertTrue(Path(roi_action["tracked_trace_xlsx"]).exists())
            self.assertEqual(len(roi_action["tracked_debug_tiffs"]), 1)
            self.assertEqual(len(roi_action["tracked_debug_raw_tiffs"]), 1)
            self.assertEqual(len(roi_action["tracked_debug_mask_tiffs"]), 1)
            self.assertTrue(Path(roi_action["tracked_debug_tiffs"][0]).exists())
            self.assertTrue(Path(roi_action["tracked_debug_raw_tiffs"][0]).exists())
            self.assertTrue(Path(roi_action["tracked_debug_mask_tiffs"][0]).exists())

            signals = record["actions"]["signals"]
            self.assertEqual(signals["signal_source"], "tracked_segmentation")
            self.assertTrue(Path(signals["xlsx"]).exists())
            self.assertTrue(Path(signals["csv"]).exists())

            comparison = record["actions"]["signal_comparison_csvs"]
            self.assertTrue(comparison["ok"])
            self.assertEqual(len(comparison["csvs"]), 2)
            self.assertFalse(comparison["variants"]["raw"].get("skipped", False))
            self.assertFalse(comparison["variants"]["tracked_segmentation"].get("skipped", False))
            self.assertTrue(comparison["variants"]["segmentation"]["skipped"])
            self.assertTrue(comparison["variants"]["motion_segmentation_on_raw"]["skipped"])
            self.assertTrue(comparison["variants"]["motion_corrected"]["skipped"])
            self.assertTrue(comparison["variants"]["motion_corrected_segmentation"]["skipped"])
            self.assertTrue(Path(comparison["xlsx"]).exists())

            workbook = load_workbook(Path(comparison["xlsx"]), data_only=True)
            try:
                self.assertEqual(workbook["Tracked Segmentation"].cell(1, 1).value, "Frame")
                self.assertEqual(workbook["Tracked Segmentation"].cell(1, 3).value, "ROI01_S01")
                self.assertEqual(workbook["Motion Corrected"].cell(2, 1).value, "Skipped")
            finally:
                workbook.close()

    def test_native_pipeline_split_writes_roi_label_map(self) -> None:
        with tempfile.TemporaryDirectory() as temp_dir:
            root = Path(temp_dir)
            tiff_path = root / "roi_sample.tiff"
            metadata_path = metadata_path_for_tiff(tiff_path)
            output_root = root / "native_pipeline"
            manifest_path = output_root / "manifest.json"
            self._write_roi_tiff(tiff_path)
            metadata_path.write_text(self._metadata_text(width=32, height=16, roi_count=2), encoding="utf-8")

            exit_code = native_pipeline.main(
                [
                    str(tiff_path),
                    "--output",
                    str(output_root),
                    "--manifest",
                    str(manifest_path),
                    "--no-average",
                    "--no-signals",
                ]
            )

            self.assertEqual(exit_code, 0)
            manifest = json.loads(manifest_path.read_text(encoding="utf-8"))
            split_action = manifest["tiffs"][0]["actions"]["split"]
            self.assertTrue(split_action["ok"])
            self.assertTrue(Path(split_action["roi_label_png"]).exists())
            with Image.open(split_action["roi_label_png"]) as label_image:
                self.assertGreater(label_image.size[0], 32)
                self.assertGreater(label_image.size[1], 16)
            split_dir = Path(split_action["output_dir"])
            self.assertTrue((split_dir / "roi_sample_ROI01.tiff").exists())
            self.assertTrue((split_dir / "roi_sample_ROI02.tiff").exists())

    def test_native_pipeline_passes_typed_roi_hook_parameters(self) -> None:
        with tempfile.TemporaryDirectory() as temp_dir:
            root = Path(temp_dir)
            tiff_path = root / "roi_sample.tiff"
            metadata_path = metadata_path_for_tiff(tiff_path)
            output_root = root / "native_pipeline"
            manifest_path = output_root / "manifest.json"
            self._write_roi_tiff(tiff_path)
            metadata_path.write_text(self._metadata_text(width=32, height=16, roi_count=2), encoding="utf-8")

            exit_code = native_pipeline.main(
                [
                    str(tiff_path),
                    "--output",
                    str(output_root),
                    "--manifest",
                    str(manifest_path),
                    "--no-average",
                    "--no-signals",
                    "--no-split",
                    "--roi-hook",
                    "image_processor.segmentation_hooks:roi_hook",
                    "--roi-hook-param",
                    "min_area=20",
                    "--roi-hook-param",
                    "percentile=90.5",
                ]
            )

            self.assertEqual(exit_code, 0)
            manifest = json.loads(manifest_path.read_text(encoding="utf-8"))
            self.assertEqual(manifest["settings"]["roi_hook_params"], {"min_area": 20, "percentile": 90.5})
            roi_action = manifest["tiffs"][0]["actions"]["roi_detection"]
            self.assertEqual(roi_action["hook_parameters"], {"min_area": 20, "percentile": 90.5})
            self.assertEqual(roi_action["detected_roi_count"], 0)

    def test_native_pipeline_segments_single_frame_after_motion_skip(self) -> None:
        with tempfile.TemporaryDirectory() as temp_dir:
            root = Path(temp_dir)
            tiff_path = root / "still_sample.tiff"
            metadata_path = metadata_path_for_tiff(tiff_path)
            output_root = root / "native_pipeline"
            manifest_path = output_root / "manifest.json"
            self._write_single_frame_tiff(tiff_path)
            metadata_path.write_text(self._metadata_text(width=32, height=16, roi_count=2), encoding="utf-8")

            exit_code = native_pipeline.main(
                [
                    str(tiff_path),
                    "--output",
                    str(output_root),
                    "--manifest",
                    str(manifest_path),
                    "--no-average",
                    "--no-signals",
                    "--no-split",
                    "--motion-hook",
                    "image_processor.motion_correction:motion_hook",
                    "--roi-hook",
                    "image_processor.segmentation_hooks:roi_hook",
                    "--roi-hook-param",
                    "min_area=4",
                ]
            )

            self.assertEqual(exit_code, 0)
            manifest = json.loads(manifest_path.read_text(encoding="utf-8"))
            record = manifest["tiffs"][0]
            self.assertTrue(record["ok"])
            self.assertEqual(Path(record["working_tiff"]), tiff_path)

            motion_action = record["actions"]["motion_correction"]
            self.assertTrue(motion_action["ok"])
            self.assertTrue(motion_action["motion_skipped"])
            self.assertEqual(motion_action["frame_count"], 1)
            self.assertNotIn("corrected_tiff", motion_action)
            self.assertTrue(Path(motion_action["qc_json"]).exists())

            roi_action = record["actions"]["roi_detection"]
            self.assertTrue(roi_action["ok"])
            self.assertEqual(roi_action["frame_count"], 1)
            self.assertEqual(roi_action["detected_roi_count"], 2)
            self.assertTrue(Path(roi_action["mask_path"]).exists())
            self.assertTrue(Path(roi_action["qc_overlay_png"]).exists())

    def test_native_pipeline_rejects_reserved_hook_parameters(self) -> None:
        with tempfile.TemporaryDirectory() as temp_dir:
            root = Path(temp_dir)
            tiff_path = root / "roi_sample.tiff"
            metadata_path = metadata_path_for_tiff(tiff_path)
            self._write_roi_tiff(tiff_path)
            metadata_path.write_text(self._metadata_text(width=32, height=16, roi_count=2), encoding="utf-8")

            exit_code = native_pipeline.main(
                [
                    str(tiff_path),
                    "--no-average",
                    "--no-signals",
                    "--no-split",
                    "--roi-hook",
                    "image_processor.segmentation_hooks:roi_hook",
                    "--roi-hook-param",
                    "tiff_path=bad",
                ]
            )

            self.assertEqual(exit_code, 2)

    def test_native_pipeline_resolves_relative_hook_outputs(self) -> None:
        original_cwd = Path.cwd()
        with tempfile.TemporaryDirectory() as temp_dir:
            root = Path(temp_dir)
            work = root / "work"
            work.mkdir()
            tiff_path = work / "roi_sample.tiff"
            metadata_path = metadata_path_for_tiff(tiff_path)
            self._write_roi_tiff(tiff_path)
            metadata_path.write_text(self._metadata_text(width=32, height=16, roi_count=2), encoding="utf-8")
            os.chdir(root)
            try:
                exit_code = native_pipeline.main(
                    [
                        "work/roi_sample.tiff",
                        "--output",
                        "work/out",
                        "--manifest",
                        "work/out/manifest.json",
                        "--sampling-rate",
                        "100",
                        "--no-average",
                        "--no-split",
                        "--motion-hook",
                        "image_processor.motion_correction:motion_hook",
                        "--roi-hook",
                        "image_processor.segmentation_hooks:roi_hook",
                    ]
                )
            finally:
                os.chdir(original_cwd)

            self.assertEqual(exit_code, 0)
            manifest = json.loads((work / "out" / "manifest.json").read_text(encoding="utf-8"))
            record = manifest["tiffs"][0]
            self.assertTrue(Path(record["working_tiff"]).exists())
            self.assertTrue(Path(record["actions"]["motion_correction"]["corrected_tiff"]).exists())
            self.assertTrue(Path(record["actions"]["roi_detection"]["mask_path"]).exists())
            self.assertTrue(Path(record["actions"]["signals"]["xlsx"]).exists())

    def test_native_pipeline_handles_relative_roi_hook_metadata(self) -> None:
        original_cwd = Path.cwd()
        with tempfile.TemporaryDirectory() as temp_dir:
            root = Path(temp_dir)
            work = root / "work"
            work.mkdir()
            tiff_path = work / "roi_sample.tiff"
            metadata_path = metadata_path_for_tiff(tiff_path)
            self._write_roi_tiff(tiff_path)
            metadata_path.write_text(self._metadata_text(width=32, height=16, roi_count=2), encoding="utf-8")
            os.chdir(root)
            try:
                exit_code = native_pipeline.main(
                    [
                        "work/roi_sample.tiff",
                        "--output",
                        "work/out",
                        "--manifest",
                        "work/out/manifest.json",
                        "--sampling-rate",
                        "100",
                        "--no-average",
                        "--no-split",
                        "--roi-hook",
                        "image_processor.segmentation_hooks:roi_hook",
                    ]
                )
            finally:
                os.chdir(original_cwd)

            self.assertEqual(exit_code, 0)
            manifest = json.loads((work / "out" / "manifest.json").read_text(encoding="utf-8"))
            record = manifest["tiffs"][0]
            self.assertTrue(record["ok"])
            self.assertTrue(Path(record["actions"]["roi_detection"]["qc_overlay_png"]).exists())
            self.assertTrue(Path(record["actions"]["signals"]["xlsx"]).exists())

    @staticmethod
    def _write_motion_tiff(path: Path) -> None:
        first = np.zeros((32, 32), dtype=np.uint16)
        first[10:15, 12:17] = 1000
        second = np.zeros((32, 32), dtype=np.uint16)
        second[12:17, 9:14] = 1000
        frames = [Image.fromarray(first), Image.fromarray(second)]
        frames[0].save(path, save_all=True, append_images=frames[1:])

    @staticmethod
    def _write_single_frame_tiff(path: Path) -> None:
        frame = np.full((16, 32), 10, dtype=np.uint16)
        frame[4:8, 5:9] = 100
        frame[9:13, 22:26] = 120
        Image.fromarray(frame).save(path)

    @staticmethod
    def _write_roi_tiff(path: Path) -> None:
        frames = []
        for frame_index in range(3):
            frame = np.full((16, 32), 10 + frame_index, dtype=np.uint16)
            frame[4:8, 5:9] = 100 + frame_index
            frame[9:13, 22:26] = 120 + frame_index
            frames.append(Image.fromarray(frame))
        frames[0].save(path, save_all=True, append_images=frames[1:])

    @staticmethod
    def _write_tracked_tiff(path: Path) -> None:
        frames = []
        for frame_index in range(4):
            frame = np.full((16, 16), 8, dtype=np.uint16)
            y = 4 + frame_index
            x = 5 + frame_index
            frame[y : y + 4, x : x + 4] = 180 + frame_index
            frames.append(Image.fromarray(frame))
        frames[0].save(path, save_all=True, append_images=frames[1:])

    @staticmethod
    def _metadata_text(*, width: int, height: int, roi_count: int) -> str:
        if roi_count == 1:
            roi_sizes = f"[[{width}, {height}]]"
            roi_indices = "[1]"
        else:
            tile_width = width // roi_count
            roi_sizes = ", ".join(f"[{tile_width}, {height}]" for _ in range(roi_count))
            roi_sizes = f"[{roi_sizes}]"
            roi_indices = "[" + ", ".join(str(index) for index in range(1, roi_count + 1)) + "]"
        return f"""%Synthetic metadata
Type: MultiROIChessBoard

>Context start
%SIZE
 Width: {width}
 Height: {height}
%PATTERN
 PatternNum: 1
 RoiCount: {roi_count}
 RoiSize (pixels): {roi_sizes}
 RoiIndex: {roi_indices}
 Channel: UG
 DataType: uint16
<Context end
"""


if __name__ == "__main__":
    unittest.main()
