"""Dynamic tracked segmentation hook for tiny Femtonics ROI tiles."""

from __future__ import annotations

import csv
import json
import math
import struct
from dataclasses import dataclass
from datetime import datetime
from pathlib import Path
from typing import Any, Iterable

import numpy as np
from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

try:
    from .roi_core import ClassicTiffStackWriter, RoiBox, metadata_path_for_tiff, parse_metadata, save_tiff_array_stack
    from .motion_correction import correct_tiff_stack, write_motion_qc
    from .segmentation_hooks import (
        _layout_boxes_for_tiff,
        _masks_from_label_image,
        detect_tile_components,
        mask_boundary,
        save_segmentation_overlay_png,
    )
    from .signal_export import resolve_sampling_rate
    from .summary_images import auto_display_range, compute_summary_images, scale_to_uint8, write_summary_outputs
    from .tiff_backend import iter_tiff_frames
except ImportError:  # pragma: no cover - supports direct script-style imports.
    from roi_core import ClassicTiffStackWriter, RoiBox, metadata_path_for_tiff, parse_metadata, save_tiff_array_stack
    from motion_correction import correct_tiff_stack, write_motion_qc
    from segmentation_hooks import (
        _layout_boxes_for_tiff,
        _masks_from_label_image,
        detect_tile_components,
        mask_boundary,
        save_segmentation_overlay_png,
    )
    from signal_export import resolve_sampling_rate
    from summary_images import auto_display_range, compute_summary_images, scale_to_uint8, write_summary_outputs
    from tiff_backend import iter_tiff_frames


PALETTE = np.asarray(
    [
        (255, 80, 80),
        (80, 220, 120),
        (80, 160, 255),
        (255, 210, 80),
        (210, 100, 255),
        (80, 230, 230),
        (255, 145, 80),
        (180, 255, 90),
    ],
    dtype=np.float32,
)


@dataclass(frozen=True)
class TrackMask:
    label: str
    label_id: int
    parent_roi_index: int
    box: RoiBox
    ys: np.ndarray
    xs: np.ndarray
    template_values: np.ndarray

    @property
    def pixel_count(self) -> int:
        return int(self.ys.size)


@dataclass
class FrameTrack:
    frame: int
    mask_index: int
    label: str
    parent_roi_index: int
    dx: int
    dy: int
    tracking_score: float
    visible_fraction: float
    visible_pixel_count: int
    clipped: bool
    edge_touch: bool
    valid_frame: bool
    trace_valid: bool
    trace_excluded_reason: str
    signal: float
    tracking_mode: str = "fixed"
    local_dx: int = 0
    local_dy: int = 0
    cumulative_dx: int = 0
    cumulative_dy: int = 0
    tracking_score_local: float = math.nan
    tracking_score_rescue: float = math.nan
    used_rescue: bool = False
    template_updated: bool = False
    track_state: str = "accepted"


class ClassicRgbTiffStackWriter:
    """Stream uint8 RGB frames to a classic multi-page TIFF file."""

    def __init__(self, output_path: Path, frame_count: int, width: int, height: int) -> None:
        if frame_count <= 0:
            raise ValueError("Cannot save an empty RGB TIFF stack.")
        if width <= 0 or height <= 0:
            raise ValueError(f"Invalid RGB TIFF frame size {width}x{height}.")
        self.output_path = output_path
        self.frame_count = int(frame_count)
        self.width = int(width)
        self.height = int(height)
        self.frame_bytes = int(width * height * 3)
        self.entries_per_frame = 10
        self.extra_bytes = 6
        self.ifd_size = 2 + self.entries_per_frame * 12 + 4 + self.extra_bytes
        total_size = 8 + self.frame_count * (self.ifd_size + self.frame_bytes)
        if total_size > 0xFFFFFFFF:
            raise ValueError("RGB debug TIFF exceeds classic TIFF size limit.")
        self.handle = None

    def __enter__(self) -> "ClassicRgbTiffStackWriter":
        self.output_path.parent.mkdir(parents=True, exist_ok=True)
        self.handle = self.output_path.open("wb")
        self.handle.write(b"II")
        self.handle.write(struct.pack("<H", 42))
        self.handle.write(struct.pack("<I", 8))
        return self

    def __exit__(self, exc_type, exc_value, traceback_obj) -> None:
        self.close()

    def close(self) -> None:
        if self.handle is not None:
            self.handle.close()
            self.handle = None

    @staticmethod
    def _short_entry(tag: int, value: int) -> bytes:
        return struct.pack("<HHI", tag, 3, 1) + struct.pack("<H", int(value)) + b"\x00\x00"

    @staticmethod
    def _long_entry(tag: int, value: int) -> bytes:
        return struct.pack("<HHII", tag, 4, 1, int(value))

    @staticmethod
    def _short_array_entry(tag: int, count: int, offset: int) -> bytes:
        return struct.pack("<HHII", tag, 3, int(count), int(offset))

    def write_frame(self, frame_number: int, frame: np.ndarray) -> None:
        if self.handle is None:
            raise RuntimeError("RGB TIFF writer is not open.")
        if frame_number < 0 or frame_number >= self.frame_count:
            raise ValueError(f"Frame number {frame_number} is outside 0..{self.frame_count - 1}.")
        array = np.asarray(frame)
        if array.shape != (self.height, self.width, 3):
            raise ValueError(f"Expected RGB frame shape {(self.height, self.width, 3)}, got {array.shape}.")
        if array.dtype != np.uint8:
            raise ValueError(f"Expected RGB uint8 frame, got {array.dtype}.")

        ifd_offset = self.handle.tell()
        bits_offset = ifd_offset + 2 + self.entries_per_frame * 12 + 4
        data_offset = ifd_offset + self.ifd_size
        next_ifd_offset = data_offset + self.frame_bytes if frame_number + 1 < self.frame_count else 0
        entries = [
            self._long_entry(256, self.width),
            self._long_entry(257, self.height),
            self._short_array_entry(258, 3, bits_offset),
            self._short_entry(259, 1),
            self._short_entry(262, 2),
            self._long_entry(273, data_offset),
            self._short_entry(277, 3),
            self._long_entry(278, self.height),
            self._long_entry(279, self.frame_bytes),
            self._short_entry(284, 1),
        ]
        self.handle.write(struct.pack("<H", len(entries)))
        for entry in entries:
            self.handle.write(entry)
        self.handle.write(struct.pack("<I", next_ifd_offset))
        self.handle.write(struct.pack("<HHH", 8, 8, 8))
        self.handle.write(np.ascontiguousarray(array).tobytes(order="C"))


def _default_output_dir(tiff_path: Path, output_dir: Path | None) -> Path:
    if output_dir is not None:
        return Path(output_dir)
    return tiff_path.parent / "exports" / tiff_path.stem


def _as_bool(value: Any, *, default: bool = False) -> bool:
    if value is None:
        return default
    if isinstance(value, bool):
        return value
    text = str(value).strip().lower()
    if text in {"1", "true", "yes", "y", "on"}:
        return True
    if text in {"0", "false", "no", "n", "off"}:
        return False
    return default


def _as_optional_float(value: Any) -> float | None:
    if value is None:
        return None
    text = str(value).strip()
    if not text or text.lower() in {"none", "null", "off", "false"}:
        return None
    return float(value)


def _initial_mask_source_mode(value: Any) -> str:
    text = str(value or "motion_corrected_mean").strip().lower().replace("-", "_")
    if text in {"raw", "raw_mean", "mean", "raw_average"}:
        return "raw_mean"
    if text in {"motion", "motion_corrected", "motion_corrected_mean", "corrected", "corrected_mean"}:
        return "motion_corrected_mean"
    raise ValueError(
        "Unknown initial_mask_source "
        f"{value!r}; expected 'motion_corrected_mean' or 'raw_mean'."
    )


def _tracking_mode(value: Any) -> str:
    text = str(value or "cumulative").strip().lower().replace("-", "_")
    if text in {"fixed", "frame_independent", "original", "origin"}:
        return "fixed"
    if text in {"cumulative", "previous", "previous_frame", "tracking"}:
        return "cumulative"
    raise ValueError(f"Unknown tracking_mode {value!r}; expected 'cumulative' or 'fixed'.")


def _resolve_max_cumulative_shift_px(
    boxes: list[RoiBox],
    *,
    local_shift_px: int,
    value: Any,
) -> int | None:
    text = "auto" if value is None else str(value).strip().lower()
    if not text or text in {"auto", "default"}:
        if not boxes:
            return int(local_shift_px)
        smallest_tile_edge = min(min(box.width, box.height) for box in boxes)
        return max(int(local_shift_px), min(4, max(1, int(smallest_tile_edge) // 3)))
    if text in {"none", "null", "off", "false", "unlimited"}:
        return None
    return int(value)


def _motion_corrected_initial_summary(
    source: Path,
    target_dir: Path,
    kwargs: dict[str, Any],
) -> tuple[Any, dict[str, Any]]:
    requested_source = _initial_mask_source_mode(kwargs.get("initial_mask_source", "motion_corrected_mean"))
    initial_info: dict[str, Any] = {
        "requested_initial_mask_source": requested_source,
        "initial_mask_source": "raw_mean",
        "initial_mask_tiff": str(source),
    }
    if requested_source == "raw_mean":
        return compute_summary_images(source), initial_info

    corrected_path = target_dir / f"{source.stem}_initial_mask_motion_corrected.tiff"
    try:
        requested_motion_shift = kwargs.get("initial_motion_max_shift_px", kwargs.get("max_shift_px", 2))
        motion_result = correct_tiff_stack(
            source,
            corrected_path,
            max_shift_px=None
            if requested_motion_shift is None or str(requested_motion_shift).strip().lower() == "auto"
            else int(requested_motion_shift),
            method=str(kwargs.get("initial_motion_method", "shared_template_residual")),
            subpixel=_as_bool(kwargs.get("initial_motion_subpixel"), default=False),
            min_peak_ratio=_as_optional_float(kwargs.get("initial_motion_min_peak_ratio")),
            high_pass_sigma=_as_optional_float(kwargs.get("initial_motion_high_pass_sigma")),
            template_iterations=int(kwargs.get("initial_motion_template_iterations", 1)),
            residual_max_shift_px=int(kwargs.get("initial_motion_residual_max_shift_px", 1)),
            residual_min_peak_ratio=_as_optional_float(
                kwargs.get("initial_motion_residual_min_peak_ratio", 1.10)
            ),
        )
        qc_path, shifts_path = write_motion_qc(
            motion_result,
            target_dir,
            f"{source.stem}_initial_mask_motion",
        )
        serializable_motion_result = dict(motion_result)
        serializable_motion_result.pop("shifts", None)
        initial_info.update(
            {
                "initial_mask_motion_result": serializable_motion_result,
                "initial_mask_motion_qc_json": str(qc_path),
            }
        )
        if shifts_path is not None:
            initial_info["initial_mask_motion_shifts_csv"] = str(shifts_path)
        corrected_tiff = motion_result.get("corrected_tiff")
        if corrected_tiff and Path(corrected_tiff).exists():
            initial_info.update(
                {
                    "initial_mask_source": "motion_corrected_mean",
                    "initial_mask_tiff": str(corrected_tiff),
                }
            )
            return compute_summary_images(Path(corrected_tiff)), initial_info
        initial_info["initial_mask_fallback_reason"] = motion_result.get(
            "reason",
            "Motion correction did not produce a corrected TIFF.",
        )
    except Exception as exc:
        initial_info["initial_mask_motion_error"] = str(exc)

    return compute_summary_images(source), initial_info


def _box_by_roi_index(boxes: list[RoiBox]) -> dict[int, RoiBox]:
    return {int(box.roi_index): box for box in boxes}


def _build_track_masks(
    *,
    label_image: np.ndarray,
    detections: list[dict[str, Any]],
    boxes: list[RoiBox],
    mean_image: np.ndarray,
) -> list[TrackMask]:
    by_roi = _box_by_roi_index(boxes)
    track_masks: list[TrackMask] = []
    for detection in detections:
        label_id = int(detection["label_id"])
        parent_roi_index = int(detection["parent_roi_index"])
        box = by_roi.get(parent_roi_index)
        if box is None:
            continue
        ys, xs = np.nonzero(label_image == label_id)
        if ys.size == 0:
            continue
        track_masks.append(
            TrackMask(
                label=str(detection["label"]),
                label_id=label_id,
                parent_roi_index=parent_roi_index,
                box=box,
                ys=ys.astype(np.int32, copy=False),
                xs=xs.astype(np.int32, copy=False),
                template_values=np.asarray(mean_image[ys, xs], dtype=np.float64),
            )
        )
    return track_masks


def _score_candidate(template_values: np.ndarray, candidate_values: np.ndarray) -> float:
    if candidate_values.size == 0:
        return -math.inf
    template = np.asarray(template_values, dtype=np.float64)
    candidate = np.asarray(candidate_values, dtype=np.float64)
    if template.size >= 2:
        template_std = float(np.std(template))
        candidate_std = float(np.std(candidate))
        if template_std > 1e-9 and candidate_std > 1e-9:
            return float(np.mean((template - np.mean(template)) * (candidate - np.mean(candidate))) / (template_std * candidate_std))
    return float(np.mean(candidate))


def _candidate_positions(mask: TrackMask, *, dx: int, dy: int) -> tuple[np.ndarray, np.ndarray, np.ndarray]:
    shifted_y = mask.ys + int(dy)
    shifted_x = mask.xs + int(dx)
    valid = (
        (shifted_y >= mask.box.upper)
        & (shifted_y < mask.box.lower)
        & (shifted_x >= mask.box.left)
        & (shifted_x < mask.box.right)
    )
    return shifted_y[valid], shifted_x[valid], valid


def _track_from_best_candidate(
    *,
    mask: TrackMask,
    frame_values: np.ndarray,
    best: tuple[float, int, int, np.ndarray, np.ndarray, np.ndarray] | None,
    center_dx: int,
    center_dy: int,
    min_tracking_score: float | None,
    min_visible_fraction: float,
    exclude_clipped_masks: bool,
    tracking_mode: str,
    track_state: str = "accepted",
    used_rescue: bool = False,
    tracking_score_local: float = math.nan,
    tracking_score_rescue: float = math.nan,
) -> FrameTrack:
    if best is None:
        return FrameTrack(
            0,
            0,
            mask.label,
            mask.parent_roi_index,
            int(center_dx),
            int(center_dy),
            -math.inf,
            0.0,
            0,
            True,
            True,
            False,
            False,
            "no_candidate",
            math.nan,
            tracking_mode=tracking_mode,
            local_dx=0,
            local_dy=0,
            cumulative_dx=int(center_dx),
            cumulative_dy=int(center_dy),
            tracking_score_local=tracking_score_local,
            tracking_score_rescue=tracking_score_rescue,
            used_rescue=used_rescue,
            track_state="invalid",
        )

    score, dx, dy, ys, xs, _valid = best
    visible_count = int(ys.size)
    visible_fraction = float(visible_count / max(mask.pixel_count, 1))
    clipped = visible_count < int(mask.pixel_count)
    edge_touch = bool(
        np.any(ys == mask.box.upper)
        or np.any(ys == mask.box.lower - 1)
        or np.any(xs == mask.box.left)
        or np.any(xs == mask.box.right - 1)
    )
    passes_score = min_tracking_score is None or score >= float(min_tracking_score)
    passes_visible = visible_fraction >= float(min_visible_fraction)
    valid_frame = bool(visible_count > 0 and passes_score and passes_visible)
    trace_excluded_reason = ""
    if not valid_frame:
        if visible_count <= 0:
            trace_excluded_reason = "no_visible_pixels"
        elif not passes_score:
            trace_excluded_reason = "tracking_score"
        elif not passes_visible:
            trace_excluded_reason = "visible_fraction"
    elif exclude_clipped_masks and clipped:
        trace_excluded_reason = "clipped_mask"
    trace_valid = bool(valid_frame and not trace_excluded_reason)
    signal = float(np.mean(frame_values[ys, xs])) if trace_valid else math.nan
    return FrameTrack(
        0,
        0,
        mask.label,
        mask.parent_roi_index,
        int(dx),
        int(dy),
        float(score),
        visible_fraction,
        visible_count,
        clipped,
        edge_touch,
        valid_frame,
        trace_valid,
        trace_excluded_reason,
        signal,
        tracking_mode=tracking_mode,
        local_dx=int(dx) - int(center_dx),
        local_dy=int(dy) - int(center_dy),
        cumulative_dx=int(dx),
        cumulative_dy=int(dy),
        tracking_score_local=tracking_score_local,
        tracking_score_rescue=tracking_score_rescue,
        used_rescue=used_rescue,
        track_state=track_state if valid_frame else "invalid",
    )


def _best_mask_shift(
    frame: np.ndarray,
    mask: TrackMask,
    *,
    center_dx: int,
    center_dy: int,
    search_radius_px: int,
    max_cumulative_shift_px: int | None = None,
) -> tuple[float, int, int, np.ndarray, np.ndarray, np.ndarray] | None:
    values = np.asarray(frame, dtype=np.float64)
    radius = max(0, int(search_radius_px))
    best: tuple[float, int, int, np.ndarray, np.ndarray, np.ndarray] | None = None
    for dy in range(int(center_dy) - radius, int(center_dy) + radius + 1):
        for dx in range(int(center_dx) - radius, int(center_dx) + radius + 1):
            if max_cumulative_shift_px is not None and (
                abs(int(dx)) > int(max_cumulative_shift_px) or abs(int(dy)) > int(max_cumulative_shift_px)
            ):
                continue
            ys, xs, valid = _candidate_positions(mask, dx=dx, dy=dy)
            if ys.size == 0:
                continue
            score = _score_candidate(mask.template_values[valid], values[ys, xs])
            if best is None or score > best[0]:
                best = (score, int(dx), int(dy), ys, xs, valid)
    return best


def estimate_mask_shift(
    frame: np.ndarray,
    mask: TrackMask,
    *,
    max_shift_px: int,
    min_tracking_score: float | None = None,
    min_visible_fraction: float = 0.0,
    exclude_clipped_masks: bool = False,
) -> FrameTrack:
    values = np.asarray(frame, dtype=np.float64)
    best = _best_mask_shift(
        values,
        mask,
        center_dx=0,
        center_dy=0,
        search_radius_px=max_shift_px,
    )
    return _track_from_best_candidate(
        mask=mask,
        frame_values=values,
        best=best,
        center_dx=0,
        center_dy=0,
        min_tracking_score=min_tracking_score,
        min_visible_fraction=min_visible_fraction,
        exclude_clipped_masks=exclude_clipped_masks,
        tracking_mode="fixed",
        tracking_score_local=best[0] if best is not None else -math.inf,
    )


def _better_track(left: FrameTrack, right: FrameTrack) -> FrameTrack:
    if left.valid_frame and not right.valid_frame:
        return left
    if right.valid_frame and not left.valid_frame:
        return right
    return left if left.tracking_score >= right.tracking_score else right


def estimate_cumulative_mask_shift(
    frame: np.ndarray,
    mask: TrackMask,
    *,
    previous_dx: int,
    previous_dy: int,
    local_shift_px: int,
    rescue_shift_px: int,
    min_tracking_score: float | None = None,
    min_visible_fraction: float = 0.25,
    max_cumulative_shift_px: int | None = None,
    exclude_clipped_masks: bool = False,
) -> FrameTrack:
    values = np.asarray(frame, dtype=np.float64)
    local_best = _best_mask_shift(
        values,
        mask,
        center_dx=previous_dx,
        center_dy=previous_dy,
        search_radius_px=local_shift_px,
        max_cumulative_shift_px=max_cumulative_shift_px,
    )
    local_score = float(local_best[0]) if local_best is not None else -math.inf
    local_track = _track_from_best_candidate(
        mask=mask,
        frame_values=values,
        best=local_best,
        center_dx=previous_dx,
        center_dy=previous_dy,
        min_tracking_score=min_tracking_score,
        min_visible_fraction=min_visible_fraction,
        exclude_clipped_masks=exclude_clipped_masks,
        tracking_mode="cumulative",
        track_state="accepted",
        tracking_score_local=local_score,
    )
    if local_track.valid_frame:
        return local_track

    rescue_candidates: list[FrameTrack] = []
    rescue_centers = [(0, 0)]
    if int(rescue_shift_px) > int(local_shift_px):
        rescue_centers.insert(0, (previous_dx, previous_dy))
    for center_dx, center_dy in rescue_centers:
        rescue_best = _best_mask_shift(
            values,
            mask,
            center_dx=center_dx,
            center_dy=center_dy,
            search_radius_px=rescue_shift_px,
            max_cumulative_shift_px=max_cumulative_shift_px,
        )
        rescue_score = float(rescue_best[0]) if rescue_best is not None else -math.inf
        rescue_candidates.append(
            _track_from_best_candidate(
                mask=mask,
                frame_values=values,
                best=rescue_best,
                center_dx=center_dx,
                center_dy=center_dy,
                min_tracking_score=min_tracking_score,
                min_visible_fraction=min_visible_fraction,
                exclude_clipped_masks=exclude_clipped_masks,
                tracking_mode="cumulative",
                track_state="rescued",
                used_rescue=True,
                tracking_score_local=local_score,
                tracking_score_rescue=rescue_score,
            )
        )
    rescue_track = rescue_candidates[0]
    for candidate in rescue_candidates[1:]:
        rescue_track = _better_track(rescue_track, candidate)
    if rescue_track.valid_frame:
        return rescue_track

    held_best = _best_mask_shift(
        values,
        mask,
        center_dx=previous_dx,
        center_dy=previous_dy,
        search_radius_px=0,
        max_cumulative_shift_px=max_cumulative_shift_px,
    )
    return _track_from_best_candidate(
        mask=mask,
        frame_values=values,
        best=held_best,
        center_dx=previous_dx,
        center_dy=previous_dy,
        min_tracking_score=None,
        min_visible_fraction=0.0,
        exclude_clipped_masks=exclude_clipped_masks,
        tracking_mode="cumulative",
        track_state="held",
        used_rescue=True,
        tracking_score_local=local_score,
        tracking_score_rescue=rescue_track.tracking_score,
    )


def _render_side_by_side_frame(
    frame: np.ndarray,
    box: RoiBox,
    tile_masks: list[tuple[TrackMask, FrameTrack]],
    *,
    black_level: float,
    white_level: float,
) -> np.ndarray:
    tile = np.asarray(frame[box.upper : box.lower, box.left : box.right], dtype=np.float32)
    base = scale_to_uint8(tile, black_level, white_level)
    left_rgb = np.repeat(base[:, :, np.newaxis], 3, axis=2).astype(np.float32)
    right_rgb = left_rgb.copy()

    for mask, track in tile_masks:
        if not track.valid_frame or track.visible_pixel_count <= 0:
            continue
        local_mask = np.zeros((box.height, box.width), dtype=bool)
        ys, xs, _valid = _candidate_positions(mask, dx=track.dx, dy=track.dy)
        local_mask[ys - box.upper, xs - box.left] = True
        if not np.any(local_mask):
            continue
        color = PALETTE[(mask.label_id - 1) % len(PALETTE)]
        boundary = mask_boundary(local_mask)
        right_rgb[local_mask] = right_rgb[local_mask] * 0.65 + color * 0.35
        right_rgb[boundary] = right_rgb[boundary] * 0.10 + color * 0.90

    side_by_side = np.concatenate([left_rgb, right_rgb], axis=1)
    return np.clip(np.rint(side_by_side), 0, 255).astype(np.uint8)


def _render_tracked_label_tile(
    box: RoiBox,
    tile_masks: list[tuple[TrackMask, FrameTrack]],
    *,
    dtype: np.dtype,
) -> np.ndarray:
    label_tile = np.zeros((box.height, box.width), dtype=dtype)
    for mask, track in tile_masks:
        if not track.valid_frame or track.visible_pixel_count <= 0:
            continue
        ys, xs, _valid = _candidate_positions(mask, dx=track.dx, dy=track.dy)
        if ys.size == 0:
            continue
        label_tile[ys - box.upper, xs - box.left] = int(mask.label_id)
    return label_tile


def _write_tracked_signals_csv(
    output_path: Path,
    *,
    time_ms: np.ndarray,
    labels: list[str],
    signals: np.ndarray,
) -> None:
    output_path.parent.mkdir(parents=True, exist_ok=True)
    with output_path.open("w", newline="", encoding="utf-8") as handle:
        writer = csv.writer(handle)
        writer.writerow(["Frame", "Time_ms", *labels])
        for frame_index in range(signals.shape[0]):
            writer.writerow(
                [
                    frame_index,
                    float(time_ms[frame_index]),
                    *[
                        "" if not np.isfinite(value) else float(value)
                        for value in signals[frame_index, :]
                    ],
                ]
            )


def _write_tracks_csv(output_path: Path, tracks: Iterable[FrameTrack]) -> None:
    output_path.parent.mkdir(parents=True, exist_ok=True)
    with output_path.open("w", newline="", encoding="utf-8") as handle:
        writer = csv.writer(handle)
        writer.writerow(
            [
                "Frame",
                "Mask",
                "Label",
                "ParentROI",
                "dx",
                "dy",
                "local_dx",
                "local_dy",
                "cumulative_dx",
                "cumulative_dy",
                "tracking_score",
                "tracking_score_local",
                "tracking_score_rescue",
                "visible_fraction",
                "visible_pixel_count",
                "clipped",
                "edge_touch",
                "valid_frame",
                "trace_valid",
                "trace_excluded_reason",
                "used_rescue",
                "template_updated",
                "tracking_mode",
                "track_state",
                "signal",
            ]
        )
        for track in tracks:
            writer.writerow(
                [
                    int(track.frame),
                    int(track.mask_index),
                    track.label,
                    int(track.parent_roi_index),
                    int(track.dx),
                    int(track.dy),
                    int(track.local_dx),
                    int(track.local_dy),
                    int(track.cumulative_dx),
                    int(track.cumulative_dy),
                    float(track.tracking_score),
                    float(track.tracking_score_local),
                    float(track.tracking_score_rescue),
                    float(track.visible_fraction),
                    int(track.visible_pixel_count),
                    int(bool(track.clipped)),
                    int(bool(track.edge_touch)),
                    int(bool(track.valid_frame)),
                    int(bool(track.trace_valid)),
                    track.trace_excluded_reason,
                    int(bool(track.used_rescue)),
                    int(bool(track.template_updated)),
                    track.tracking_mode,
                    track.track_state,
                    "" if not np.isfinite(track.signal) else float(track.signal),
                ]
            )


def _write_tracked_signals_xlsx(
    output_path: Path,
    *,
    time_ms: np.ndarray,
    labels: list[str],
    signals: np.ndarray,
    tracks_csv: Path,
    manifest_path: Path,
) -> None:
    output_path.parent.mkdir(parents=True, exist_ok=True)
    workbook = Workbook()
    header_fill = PatternFill("solid", fgColor="1F4E78")
    header_font = Font(bold=True, color="FFFFFF")

    sheet = workbook.active
    sheet.title = "Tracked Signals"
    sheet.append(["Frame", "Time_ms", *labels])
    for frame_index in range(signals.shape[0]):
        sheet.append(
            [
                frame_index,
                float(time_ms[frame_index]),
                *[
                    None if not np.isfinite(value) else float(value)
                    for value in signals[frame_index, :]
                ],
            ]
        )
    for cell in sheet[1]:
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center")
    sheet.freeze_panes = "C2"
    sheet.auto_filter.ref = sheet.dimensions
    sheet.column_dimensions["A"].width = 12
    sheet.column_dimensions["B"].width = 14
    for col_index in range(3, len(labels) + 3):
        sheet.column_dimensions[get_column_letter(col_index)].width = 16
    for row in sheet.iter_rows(min_row=2, min_col=2, max_col=len(labels) + 2):
        for cell in row:
            cell.number_format = "0.000"

    metadata = workbook.create_sheet("Tracked Metadata")
    metadata.append(["Field", "Value"])
    metadata.append(["Signal calculation", "Mean raw pixel intensity inside visible tracked segmentation mask pixels"])
    metadata.append(["Excluded frames", "Blank only when the tracked mask fails trace QC"])
    metadata.append(["Tracks CSV", str(tracks_csv)])
    metadata.append(["Manifest", str(manifest_path)])
    metadata.append(["Mask count", len(labels)])
    for cell in metadata[1]:
        cell.fill = header_fill
        cell.font = header_font
    metadata.column_dimensions["A"].width = 28
    metadata.column_dimensions["B"].width = 96
    workbook.save(output_path)


def _track_masks_and_write_debug(
    *,
    tiff_path: Path,
    output_dir: Path,
    boxes: list[RoiBox],
    track_masks: list[TrackMask],
    frame_count: int,
    sampling_rate_hz: float,
    max_shift_px: int,
    min_tracking_score: float | None,
    debug_tiles: bool = True,
    tracking_mode: str = "cumulative",
    local_shift_px: int | None = None,
    rescue_shift_px: int | None = None,
    min_visible_fraction: float = 0.25,
    max_cumulative_shift_px: int | None = None,
    exclude_clipped_masks: bool = False,
    progress_interval: int = 500,
) -> dict[str, Any]:
    normalized_tracking_mode = _tracking_mode(tracking_mode)
    local_radius = int(max_shift_px if local_shift_px is None else local_shift_px)
    rescue_radius = int(local_radius if rescue_shift_px is None else rescue_shift_px)
    labels = [mask.label for mask in track_masks]
    signals = np.full((frame_count, len(track_masks)), np.nan, dtype=np.float64)
    tracks: list[FrameTrack] = []
    time_ms = np.arange(frame_count, dtype=np.float64) * 1000.0 / float(sampling_rate_hz)
    cumulative_positions = {
        index: (0, 0)
        for index in range(len(track_masks))
    }

    tile_ranges: dict[int, tuple[float, float]] = {}
    writers: dict[int, ClassicRgbTiffStackWriter] = {}
    raw_writers: dict[int, ClassicTiffStackWriter] = {}
    mask_writers: dict[int, ClassicTiffStackWriter] = {}
    debug_paths: dict[int, Path] = {}
    raw_paths: dict[int, Path] = {}
    mask_paths: dict[int, Path] = {}
    mask_dtype = np.dtype(np.uint16 if len(track_masks) <= np.iinfo(np.uint16).max else np.uint32)
    try:
        if debug_tiles:
            for box in boxes:
                debug_path = output_dir / f"{tiff_path.stem}_ROI{box.roi_index:02d}_tracked_debug.tiff"
                writer = ClassicRgbTiffStackWriter(debug_path, frame_count, box.width * 2, box.height)
                writers[box.roi_index] = writer.__enter__()
                debug_paths[box.roi_index] = debug_path

                raw_path = output_dir / f"{tiff_path.stem}_ROI{box.roi_index:02d}_tracked_raw.tiff"
                mask_path = output_dir / f"{tiff_path.stem}_ROI{box.roi_index:02d}_tracked_mask.tiff"
                raw_paths[box.roi_index] = raw_path
                mask_paths[box.roi_index] = mask_path

        for frame_index, frame in enumerate(iter_tiff_frames(tiff_path)):
            if frame_index >= frame_count:
                break
            values = np.asarray(frame)
            if debug_tiles and not raw_writers:
                for box in boxes:
                    raw_writer = ClassicTiffStackWriter(
                        raw_paths[box.roi_index],
                        frame_count,
                        box.width,
                        box.height,
                        np.dtype(values.dtype),
                    )
                    mask_writer = ClassicTiffStackWriter(
                        mask_paths[box.roi_index],
                        frame_count,
                        box.width,
                        box.height,
                        mask_dtype,
                    )
                    raw_writers[box.roi_index] = raw_writer.__enter__()
                    mask_writers[box.roi_index] = mask_writer.__enter__()
            frame_tracks_by_roi: dict[int, list[tuple[TrackMask, FrameTrack]]] = {box.roi_index: [] for box in boxes}
            for mask_index, mask in enumerate(track_masks):
                if normalized_tracking_mode == "fixed":
                    track = estimate_mask_shift(
                        values,
                        mask,
                        max_shift_px=max_shift_px,
                        min_tracking_score=min_tracking_score,
                        min_visible_fraction=min_visible_fraction,
                        exclude_clipped_masks=exclude_clipped_masks,
                    )
                else:
                    previous_dx, previous_dy = cumulative_positions[mask_index]
                    track = estimate_cumulative_mask_shift(
                        values,
                        mask,
                        previous_dx=previous_dx,
                        previous_dy=previous_dy,
                        local_shift_px=local_radius,
                        rescue_shift_px=rescue_radius,
                        min_tracking_score=min_tracking_score,
                        min_visible_fraction=min_visible_fraction,
                        max_cumulative_shift_px=max_cumulative_shift_px,
                        exclude_clipped_masks=exclude_clipped_masks,
                    )
                    if track.track_state in {"accepted", "rescued", "held"} and track.valid_frame:
                        cumulative_positions[mask_index] = (track.cumulative_dx, track.cumulative_dy)
                track.frame = int(frame_index)
                track.mask_index = int(mask_index + 1)
                tracks.append(track)
                signals[frame_index, mask_index] = track.signal
                frame_tracks_by_roi.setdefault(mask.parent_roi_index, []).append((mask, track))

            for box in boxes:
                raw_writer = raw_writers.get(box.roi_index)
                if raw_writer is not None:
                    raw_writer.write_frame(frame_index, values[box.upper : box.lower, box.left : box.right])
                mask_writer = mask_writers.get(box.roi_index)
                if mask_writer is not None:
                    mask_tile = _render_tracked_label_tile(
                        box,
                        frame_tracks_by_roi.get(box.roi_index, []),
                        dtype=mask_dtype,
                    )
                    mask_writer.write_frame(frame_index, mask_tile)

                writer = writers.get(box.roi_index)
                if writer is None:
                    continue
                if box.roi_index not in tile_ranges:
                    tile = values[box.upper : box.lower, box.left : box.right]
                    tile_ranges[box.roi_index] = auto_display_range(np.asarray(tile, dtype=np.float32))
                low, high = tile_ranges[box.roi_index]
                debug_frame = _render_side_by_side_frame(
                    values,
                    box,
                    frame_tracks_by_roi.get(box.roi_index, []),
                    black_level=low,
                    white_level=high,
                )
                writer.write_frame(frame_index, debug_frame)

            if frame_index == 0 or (frame_index + 1) % progress_interval == 0 or frame_index + 1 == frame_count:
                print(f"{tiff_path.name}: tracked segmentation frames {frame_index + 1}/{frame_count}")
    finally:
        for writer in writers.values():
            writer.close()
        for writer in raw_writers.values():
            writer.close()
        for writer in mask_writers.values():
            writer.close()

    tracks_csv = output_dir / f"{tiff_path.stem}_tracked_segmentation_tracks.csv"
    _write_tracks_csv(tracks_csv, tracks)
    trace_csv = output_dir / f"{tiff_path.stem}_tracked_segmentation_signals.csv"
    _write_tracked_signals_csv(trace_csv, time_ms=time_ms, labels=labels, signals=signals)

    return {
        "labels": labels,
        "signals": signals,
        "time_ms": time_ms,
        "tracks": tracks,
        "tracks_csv": tracks_csv,
        "trace_csv": trace_csv,
        "debug_tiffs": [debug_paths[index] for index in sorted(debug_paths)],
        "debug_raw_tiffs": [raw_paths[index] for index in sorted(raw_paths)],
        "debug_mask_tiffs": [mask_paths[index] for index in sorted(mask_paths)],
        "tracking_mode": normalized_tracking_mode,
        "local_shift_px": local_radius,
        "rescue_shift_px": rescue_radius,
        "min_visible_fraction": float(min_visible_fraction),
        "max_cumulative_shift_px": max_cumulative_shift_px,
        "exclude_clipped_masks": bool(exclude_clipped_masks),
    }


def roi_hook(
    tiff_path: Path,
    output_dir: Path | None = None,
    **kwargs: Any,
) -> dict[str, Any]:
    source = Path(tiff_path)
    target_dir = _default_output_dir(source, Path(output_dir) if output_dir is not None else None)
    target_dir.mkdir(parents=True, exist_ok=True)

    summary, initial_mask_info = _motion_corrected_initial_summary(source, target_dir, kwargs)
    summary_paths = write_summary_outputs(summary, target_dir, tiff_stem=source.stem)
    boxes, metadata_path, layout_source = _layout_boxes_for_tiff(source, summary.mean.shape)
    label_image, detections = detect_tile_components(
        summary.mean,
        boxes,
        min_area=int(kwargs.get("min_area", 4)),
        percentile=float(kwargs.get("percentile", 95.0)),
        foreground_percentile=float(kwargs.get("foreground_percentile", 60.0)),
        std_factor=float(kwargs.get("std_factor", 2.5)),
        max_area_fraction=float(kwargs.get("max_area_fraction", 0.6)),
        threshold_method=str(kwargs.get("threshold_method", "foreground")),
    )

    mask_path = target_dir / f"{source.stem}_segmentation_masks.npz"
    labels = np.asarray([detection["label"] for detection in detections], dtype=str)
    np.savez_compressed(
        mask_path,
        masks=_masks_from_label_image(label_image, detections),
        label_image=label_image,
        labels=labels,
    )

    label_tiff_path = target_dir / f"{source.stem}_segmentation_labels.tiff"
    label_dtype = np.uint16 if len(detections) <= np.iinfo(np.uint16).max else np.uint32
    save_tiff_array_stack(label_tiff_path, label_image.astype(label_dtype, copy=False)[np.newaxis, :, :])
    overlay_path = save_segmentation_overlay_png(
        summary.mean,
        label_image,
        boxes,
        detections,
        target_dir / f"{source.stem}_segmentation_overlay.png",
    )

    metadata = parse_metadata(metadata_path_for_tiff(source))
    sampling_rate_hz = resolve_sampling_rate(metadata, kwargs.get("sampling_rate_hz"))
    track_masks = _build_track_masks(
        label_image=label_image,
        detections=detections,
        boxes=boxes,
        mean_image=summary.mean,
    )
    tracking_mode = _tracking_mode(kwargs.get("tracking_mode", "cumulative"))
    local_shift_px = int(kwargs.get("local_shift_px", kwargs.get("max_shift_px", 2)))
    rescue_shift_px = int(kwargs.get("rescue_shift_px", local_shift_px))
    min_visible_fraction = float(kwargs.get("min_visible_fraction", 0.25))
    exclude_clipped_masks = _as_bool(kwargs.get("exclude_clipped_masks"), default=False)
    max_cumulative_shift_px = _resolve_max_cumulative_shift_px(
        boxes,
        local_shift_px=local_shift_px,
        value=kwargs.get("max_cumulative_shift_px", "auto"),
    )
    tracking = _track_masks_and_write_debug(
        tiff_path=source,
        output_dir=target_dir,
        boxes=boxes,
        track_masks=track_masks,
        frame_count=summary.frame_count,
        sampling_rate_hz=sampling_rate_hz,
        max_shift_px=int(kwargs.get("max_shift_px", 2)),
        min_tracking_score=_as_optional_float(kwargs.get("min_tracking_score")),
        tracking_mode=tracking_mode,
        local_shift_px=local_shift_px,
        rescue_shift_px=rescue_shift_px,
        min_visible_fraction=min_visible_fraction,
        max_cumulative_shift_px=max_cumulative_shift_px,
        exclude_clipped_masks=exclude_clipped_masks,
        debug_tiles=_as_bool(kwargs.get("debug_tiles"), default=True),
    )

    manifest_path = target_dir / f"{source.stem}_tracked_segmentation_manifest.json"
    manifest = {
        "ok": True,
        "mode": "tracked_segmentation",
        "source_tiff": str(source),
        **initial_mask_info,
        "metadata_path": str(metadata_path),
        "frame_count": int(summary.frame_count),
        "sampling_rate_hz": float(sampling_rate_hz),
        "layout_source": layout_source,
        "tile_count": len(boxes),
        "detected_roi_count": len(detections),
        "tracked_mask_count": len(track_masks),
        "static_mask_path": str(mask_path),
        "static_label_tiff": str(label_tiff_path),
        "static_overlay_png": str(overlay_path),
        "summary_outputs": {key: str(path) for key, path in summary_paths.items()},
        "tracked_trace_csv": str(tracking["trace_csv"]),
        "tracked_tracks_csv": str(tracking["tracks_csv"]),
        "tracked_debug_tiffs": [str(path) for path in tracking["debug_tiffs"]],
        "tracked_debug_raw_tiffs": [str(path) for path in tracking["debug_raw_tiffs"]],
        "tracked_debug_mask_tiffs": [str(path) for path in tracking["debug_mask_tiffs"]],
        "parameters": {
            "max_shift_px": int(kwargs.get("max_shift_px", 2)),
            "tracking_mode": tracking["tracking_mode"],
            "local_shift_px": tracking["local_shift_px"],
            "rescue_shift_px": tracking["rescue_shift_px"],
            "min_tracking_score": _as_optional_float(kwargs.get("min_tracking_score")),
            "min_visible_fraction": tracking["min_visible_fraction"],
            "max_cumulative_shift_px": tracking["max_cumulative_shift_px"],
            "exclude_clipped_masks": tracking["exclude_clipped_masks"],
            "initial_mask_source": initial_mask_info.get("initial_mask_source"),
            "requested_initial_mask_source": initial_mask_info.get("requested_initial_mask_source"),
            "initial_motion_method": str(kwargs.get("initial_motion_method", "shared_template_residual")),
            "initial_motion_max_shift_px": kwargs.get("initial_motion_max_shift_px", kwargs.get("max_shift_px", 2)),
            "initial_motion_residual_max_shift_px": int(kwargs.get("initial_motion_residual_max_shift_px", 1)),
            "initial_motion_residual_min_peak_ratio": kwargs.get("initial_motion_residual_min_peak_ratio", 1.10),
            "trace_mode": "visible_tracked_mean",
            "debug_movie": "per_roi_side_by_side_tiff",
            "threshold_method": str(kwargs.get("threshold_method", "foreground")),
            "min_area": int(kwargs.get("min_area", 4)),
        },
        "detections": detections,
        "track_masks": [
            {
                "label": mask.label,
                "label_id": int(mask.label_id),
                "parent_roi_index": int(mask.parent_roi_index),
                "pixel_count": int(mask.pixel_count),
            }
            for mask in track_masks
        ],
        "processed_at": datetime.now().isoformat(timespec="seconds"),
    }
    manifest_path.write_text(json.dumps(manifest, indent=2), encoding="utf-8")

    trace_xlsx = target_dir / f"{source.stem}_tracked_segmentation_signals.xlsx"
    _write_tracked_signals_xlsx(
        trace_xlsx,
        time_ms=tracking["time_ms"],
        labels=tracking["labels"],
        signals=tracking["signals"],
        tracks_csv=tracking["tracks_csv"],
        manifest_path=manifest_path,
    )
    manifest["tracked_trace_xlsx"] = str(trace_xlsx)
    manifest_path.write_text(json.dumps(manifest, indent=2), encoding="utf-8")

    return {
        "ok": True,
        "mode": "tracked_segmentation",
        "metadata_path": str(metadata_path),
        **initial_mask_info,
        "detected_roi_count": len(detections),
        "tracked_mask_count": len(track_masks),
        "mask_path": str(mask_path),
        "mask_manifest_path": str(manifest_path),
        "label_tiff": str(label_tiff_path),
        "qc_overlay_png": str(overlay_path),
        "tracked_trace_csv": str(tracking["trace_csv"]),
        "tracked_trace_xlsx": str(trace_xlsx),
        "tracked_manifest_path": str(manifest_path),
        "tracked_tracks_csv": str(tracking["tracks_csv"]),
        "tracked_debug_tiffs": [str(path) for path in tracking["debug_tiffs"]],
        "tracked_debug_raw_tiffs": [str(path) for path in tracking["debug_raw_tiffs"]],
        "tracked_debug_mask_tiffs": [str(path) for path in tracking["debug_mask_tiffs"]],
        "tracking_mode": tracking["tracking_mode"],
        "local_shift_px": tracking["local_shift_px"],
        "rescue_shift_px": tracking["rescue_shift_px"],
        "min_visible_fraction": tracking["min_visible_fraction"],
        "max_cumulative_shift_px": tracking["max_cumulative_shift_px"],
        "exclude_clipped_masks": tracking["exclude_clipped_masks"],
        "sampling_rate_hz": float(sampling_rate_hz),
        "frame_count": int(summary.frame_count),
        "layout_source": layout_source,
        "summary_paths": [str(path) for path in summary_paths.values()],
        **{key: str(path) for key, path in summary_paths.items()},
    }


run = roi_hook
