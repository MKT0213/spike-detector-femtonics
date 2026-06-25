"""Mask-aware signal extraction and document export for TIFF stacks."""

from __future__ import annotations

import csv
import json
from dataclasses import dataclass
from datetime import datetime
from pathlib import Path
from typing import Callable, Iterable

import numpy as np
from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

try:
    from .roi_core import (
        RoiMetadata,
        TiffInfo,
        compute_roi_layout,
        load_tiff_info,
        metadata_path_for_tiff,
        parse_metadata,
        validate_tiff_matches_metadata,
    )
    from .signal_export import default_export_output_dir, resolve_sampling_rate
    from .tiff_backend import iter_tiff_frames
except ImportError:  # pragma: no cover - supports direct script-style imports.
    from roi_core import (
        RoiMetadata,
        TiffInfo,
        compute_roi_layout,
        load_tiff_info,
        metadata_path_for_tiff,
        parse_metadata,
        validate_tiff_matches_metadata,
    )
    from signal_export import default_export_output_dir, resolve_sampling_rate
    from tiff_backend import iter_tiff_frames


MaskSignalProgress = Callable[[int, int], None]


@dataclass(frozen=True)
class MaskRecord:
    label: str
    mask_index: int
    pixel_count: int
    weight_sum: float
    bbox: tuple[int, int, int, int] | None
    centroid: tuple[float, float] | None
    parent_roi_index: int | None = None


@dataclass(frozen=True)
class MaskSignalTable:
    source_tiff: Path
    metadata_path: Path
    metadata: RoiMetadata
    tiff_info: TiffInfo
    mask_path: Path
    mask_manifest_path: Path | None
    mask_records: list[MaskRecord]
    sampling_rate_hz: float
    time_ms: np.ndarray
    signals: np.ndarray
    raw_signals: np.ndarray
    background_signals: np.ndarray | None
    extraction_mode: str
    background_mode: str
    processed_at: str

    @property
    def roi_labels(self) -> list[str]:
        return [record.label for record in self.mask_records]

    @property
    def headers(self) -> list[str]:
        return ["Frame", "Time_ms", *self.roi_labels]

    def iter_signal_rows(self) -> Iterable[list[float | int]]:
        for frame_index in range(self.tiff_info.frame_count):
            yield [
                frame_index,
                float(self.time_ms[frame_index]),
                *[float(value) for value in self.signals[frame_index, :]],
            ]


@dataclass(frozen=True)
class MaskExportResult:
    signal_table: MaskSignalTable
    xlsx_path: Path
    csv_path: Path | None


def _coerce_labels(raw_labels: np.ndarray | None, mask_count: int) -> list[str]:
    if raw_labels is None or len(raw_labels) != mask_count:
        return [f"MASK{index:02d}" for index in range(1, mask_count + 1)]

    labels: list[str] = []
    for index, value in enumerate(raw_labels.tolist(), start=1):
        if isinstance(value, bytes):
            text = value.decode("utf-8", errors="replace")
        else:
            text = str(value)
        labels.append(text or f"MASK{index:02d}")
    return labels


def load_mask_array(mask_path: Path) -> tuple[np.ndarray, list[str]]:
    with np.load(mask_path, allow_pickle=False) as data:
        if "masks" in data:
            masks = np.asarray(data["masks"])
        elif "label_image" in data:
            label_image = np.asarray(data["label_image"])
            max_label = int(np.max(label_image)) if label_image.size else 0
            if max_label <= 0:
                masks = np.zeros((0, *label_image.shape), dtype=bool)
            else:
                labels_present = [label for label in range(1, max_label + 1)]
                masks = np.asarray([(label_image == label) for label in labels_present], dtype=bool)
        else:
            raise ValueError(f"Mask file must contain 'masks' or 'label_image': {mask_path}")

        raw_labels = np.asarray(data["labels"]) if "labels" in data else None

    if masks.ndim == 2:
        masks = masks[np.newaxis, :, :]
    if masks.ndim != 3:
        raise ValueError(f"Masks must have shape (N, height, width), got {masks.shape}.")
    return masks, _coerce_labels(raw_labels, int(masks.shape[0]))


def normalize_mask_weights(masks: np.ndarray) -> tuple[np.ndarray, str]:
    values = np.asarray(masks)
    if values.shape[0] == 0:
        return values.astype(np.float64), "empty_mask_set"

    if values.dtype == np.bool_:
        return values.astype(np.float64), "binary_mean"

    finite = values[np.isfinite(values)]
    if finite.size == 0:
        raise ValueError("Mask array contains no finite values.")

    unique = np.unique(finite)
    if unique.size <= 2 and np.all(np.isin(unique, [0, 1])):
        return (values != 0).astype(np.float64), "binary_mean"

    weights = values.astype(np.float64, copy=False)
    if np.any(weights < 0):
        raise ValueError("Weighted masks must not contain negative values.")
    return weights, "weighted_mean"


def _mask_manifest_by_label(mask_manifest_path: Path | None) -> dict[str, dict]:
    if mask_manifest_path is None or not mask_manifest_path.exists():
        return {}
    try:
        manifest = json.loads(mask_manifest_path.read_text(encoding="utf-8"))
    except (OSError, json.JSONDecodeError):
        return {}
    detections = manifest.get("detections", [])
    if not isinstance(detections, list):
        return {}
    return {
        str(detection.get("label")): detection
        for detection in detections
        if isinstance(detection, dict) and detection.get("label")
    }


def build_mask_records(
    weights: np.ndarray,
    labels: list[str],
    mask_manifest_path: Path | None,
) -> list[MaskRecord]:
    by_label = _mask_manifest_by_label(mask_manifest_path)
    records: list[MaskRecord] = []
    for index, label in enumerate(labels):
        mask = np.asarray(weights[index])
        positive = mask > 0
        pixel_count = int(np.count_nonzero(positive))
        weight_sum = float(np.sum(mask))
        if pixel_count == 0 or weight_sum <= 0:
            raise ValueError(f"Mask {label} contains no positive pixels.")

        ys, xs = np.nonzero(positive)
        bbox = (
            int(np.min(xs)),
            int(np.min(ys)),
            int(np.max(xs)) + 1,
            int(np.max(ys)) + 1,
        )
        centroid = (
            float(np.average(xs, weights=mask[positive])),
            float(np.average(ys, weights=mask[positive])),
        )
        detection = by_label.get(label, {})
        parent_roi_index = detection.get("parent_roi_index")
        records.append(
            MaskRecord(
                label=label,
                mask_index=index + 1,
                pixel_count=pixel_count,
                weight_sum=weight_sum,
                bbox=bbox,
                centroid=centroid,
                parent_roi_index=int(parent_roi_index) if parent_roi_index is not None else None,
            )
        )
    return records


def normalize_background_mode(background_mode: str | None) -> str:
    mode = (background_mode or "none").strip().lower().replace("-", "_")
    aliases = {
        "": "none",
        "off": "none",
        "false": "none",
        "raw": "none",
        "tile": "parent_tile",
        "parent": "parent_tile",
        "roi_tile": "parent_tile",
        "parent_roi": "parent_tile",
        "whole_frame": "global",
        "frame": "global",
    }
    mode = aliases.get(mode, mode)
    if mode not in {"none", "parent_tile", "global"}:
        raise ValueError(
            f"Unsupported mask background mode {background_mode!r}; "
            "expected none, parent_tile, or global."
        )
    return mode


def build_background_weights(
    *,
    weights: np.ndarray,
    mask_records: list[MaskRecord],
    metadata: RoiMetadata,
    background_mode: str,
) -> tuple[np.ndarray | None, np.ndarray | None]:
    mode = normalize_background_mode(background_mode)
    if mode == "none" or not mask_records:
        return None, None

    height, width = weights.shape[1:]
    mask_union = np.any(weights > 0, axis=0)
    background_weights = np.zeros_like(weights, dtype=np.float64)
    layout_by_roi_index = {}
    if mode == "parent_tile":
        layout = compute_roi_layout(metadata)
        layout_by_roi_index = {box.roi_index: box for box in layout.boxes}

    for index, record in enumerate(mask_records):
        candidate = np.ones((height, width), dtype=bool)
        if mode == "parent_tile" and record.parent_roi_index is not None:
            box = layout_by_roi_index.get(record.parent_roi_index)
            if box is not None:
                candidate[:, :] = False
                candidate[box.upper : box.lower, box.left : box.right] = True
        candidate &= ~mask_union
        pixel_count = int(np.count_nonzero(candidate))
        if pixel_count <= 0:
            raise ValueError(
                f"Background mode {mode} found no non-mask background pixels for {record.label}."
            )
        background_weights[index, candidate] = 1.0

    sums = np.sum(background_weights, axis=(1, 2))
    return background_weights, sums.astype(np.float64)


def extract_mask_signals(
    tiff_path: Path,
    mask_path: Path,
    *,
    mask_manifest_path: Path | None = None,
    sampling_rate_hz: float | None = None,
    background_mode: str = "none",
    progress: MaskSignalProgress | None = None,
) -> MaskSignalTable:
    metadata_path = metadata_path_for_tiff(tiff_path)
    if not metadata_path.exists():
        raise FileNotFoundError(f"Metadata file is missing: {metadata_path}")

    metadata = parse_metadata(metadata_path)
    rate = resolve_sampling_rate(metadata, sampling_rate_hz)
    tiff_info = load_tiff_info(tiff_path)
    validate_tiff_matches_metadata(tiff_info, metadata)

    masks, labels = load_mask_array(mask_path)
    weights, extraction_mode = normalize_mask_weights(masks)
    if weights.shape[1:] != (tiff_info.frame_height, tiff_info.frame_width):
        raise ValueError(
            f"Mask shape {weights.shape[1:]} does not match TIFF frame shape "
            f"{(tiff_info.frame_height, tiff_info.frame_width)}."
        )

    mask_records = build_mask_records(weights, labels, mask_manifest_path) if weights.shape[0] else []
    weight_sums = np.asarray([record.weight_sum for record in mask_records], dtype=np.float64)
    normalized_background_mode = normalize_background_mode(background_mode)
    background_weights, background_weight_sums = build_background_weights(
        weights=weights,
        mask_records=mask_records,
        metadata=metadata,
        background_mode=normalized_background_mode,
    )
    raw_signals = np.zeros((tiff_info.frame_count, len(mask_records)), dtype=np.float64)
    background_signals = (
        np.zeros((tiff_info.frame_count, len(mask_records)), dtype=np.float64)
        if background_weights is not None and background_weight_sums is not None
        else None
    )

    for frame_number, frame in enumerate(iter_tiff_frames(tiff_path)):
        if frame_number >= tiff_info.frame_count:
            break
        values = np.asarray(frame, dtype=np.float64)
        if values.shape != weights.shape[1:]:
            raise ValueError(f"Frame {frame_number} shape {values.shape} does not match mask shape {weights.shape[1:]}.")
        if len(mask_records):
            raw_signals[frame_number, :] = np.tensordot(weights, values, axes=([1, 2], [0, 1])) / weight_sums
            if background_weights is not None and background_weight_sums is not None and background_signals is not None:
                background_signals[frame_number, :] = (
                    np.tensordot(background_weights, values, axes=([1, 2], [0, 1])) / background_weight_sums
                )

        if progress is not None and (
            frame_number == 0
            or (frame_number + 1) % 500 == 0
            or frame_number + 1 == tiff_info.frame_count
        ):
            progress(frame_number + 1, tiff_info.frame_count)

    frame_index = np.arange(tiff_info.frame_count, dtype=np.float64)
    time_ms = frame_index * 1000.0 / rate
    signals = raw_signals if background_signals is None else raw_signals - background_signals
    return MaskSignalTable(
        source_tiff=tiff_path,
        metadata_path=metadata_path,
        metadata=metadata,
        tiff_info=tiff_info,
        mask_path=mask_path,
        mask_manifest_path=mask_manifest_path,
        mask_records=mask_records,
        sampling_rate_hz=rate,
        time_ms=time_ms,
        signals=signals,
        raw_signals=raw_signals,
        background_signals=background_signals,
        extraction_mode=extraction_mode,
        background_mode=normalized_background_mode,
        processed_at=datetime.now().isoformat(timespec="seconds"),
    )


def write_mask_signals_csv(signal_table: MaskSignalTable, output_path: Path) -> None:
    output_path.parent.mkdir(parents=True, exist_ok=True)
    with output_path.open("w", newline="", encoding="utf-8") as handle:
        writer = csv.writer(handle)
        writer.writerow(signal_table.headers)
        writer.writerows(signal_table.iter_signal_rows())


def metadata_sheet_rows(signal_table: MaskSignalTable) -> list[tuple[str, str | int | float | None]]:
    metadata = signal_table.metadata
    info = signal_table.tiff_info
    return [
        ("Source TIFF", str(signal_table.source_tiff)),
        ("Metadata file", str(signal_table.metadata_path)),
        ("Mask file", str(signal_table.mask_path)),
        ("Mask manifest", str(signal_table.mask_manifest_path) if signal_table.mask_manifest_path else None),
        ("Processed at", signal_table.processed_at),
        ("Signal calculation", "Mean raw pixel intensity inside segmentation masks"),
        ("Extraction mode", signal_table.extraction_mode),
        ("Background mode", signal_table.background_mode),
        ("Sampling rate Hz", signal_table.sampling_rate_hz),
        ("Time unit", "ms"),
        ("Frame count", info.frame_count),
        ("Frame width", info.frame_width),
        ("Frame height", info.frame_height),
        ("TIFF mode", info.mode),
        ("TIFF dtype", info.dtype),
        ("Acquisition type", metadata.acquisition_type),
        ("Comment", metadata.comment),
        ("Measurement date", metadata.measurement_date),
        ("Channel", metadata.channel),
        ("Metadata data type", metadata.data_type),
        ("Mask count", len(signal_table.mask_records)),
    ]


def write_mask_signals_xlsx(signal_table: MaskSignalTable, output_path: Path) -> None:
    output_path.parent.mkdir(parents=True, exist_ok=True)
    workbook = Workbook()
    signals_sheet = workbook.active
    signals_sheet.title = "Signals"

    header_fill = PatternFill("solid", fgColor="1F4E78")
    header_font = Font(bold=True, color="FFFFFF")
    subtle_fill = PatternFill("solid", fgColor="EAF2F8")

    signals_sheet.append(signal_table.headers)
    for row in signal_table.iter_signal_rows():
        signals_sheet.append(row)

    for cell in signals_sheet[1]:
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center")

    signals_sheet.freeze_panes = "C2"
    signals_sheet.auto_filter.ref = signals_sheet.dimensions
    signals_sheet.column_dimensions["A"].width = 12
    signals_sheet.column_dimensions["B"].width = 14
    for col_index in range(3, len(signal_table.headers) + 1):
        signals_sheet.column_dimensions[get_column_letter(col_index)].width = 16
    for row in signals_sheet.iter_rows(min_row=2, min_col=2, max_col=len(signal_table.headers)):
        for cell in row:
            cell.number_format = "0.000"

    if signal_table.background_signals is not None:
        background_sheet = workbook.create_sheet("Background")
        background_sheet.append(signal_table.headers)
        for frame_index in range(signal_table.tiff_info.frame_count):
            background_sheet.append(
                [
                    frame_index,
                    float(signal_table.time_ms[frame_index]),
                    *[float(value) for value in signal_table.background_signals[frame_index, :]],
                ]
            )
        for cell in background_sheet[1]:
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal="center")
        background_sheet.freeze_panes = "C2"
        background_sheet.auto_filter.ref = background_sheet.dimensions
        background_sheet.column_dimensions["A"].width = 12
        background_sheet.column_dimensions["B"].width = 14
        for col_index in range(3, len(signal_table.headers) + 1):
            background_sheet.column_dimensions[get_column_letter(col_index)].width = 16
        for row in background_sheet.iter_rows(min_row=2, min_col=2, max_col=len(signal_table.headers)):
            for cell in row:
                cell.number_format = "0.000"

    metadata_sheet = workbook.create_sheet("Metadata")
    metadata_sheet.append(["Field", "Value"])
    for row in metadata_sheet_rows(signal_table):
        metadata_sheet.append(list(row))
    for cell in metadata_sheet[1]:
        cell.fill = header_fill
        cell.font = header_font
    metadata_sheet.column_dimensions["A"].width = 24
    metadata_sheet.column_dimensions["B"].width = 96
    metadata_sheet.freeze_panes = "A2"

    mask_sheet = workbook.create_sheet("Mask_Map")
    mask_sheet.append(
        [
            "ROI",
            "Mask_Index",
            "Parent_ROI",
            "Pixels",
            "Weight_Sum",
            "Left",
            "Upper",
            "Right",
            "Lower",
            "Centroid_X",
            "Centroid_Y",
        ]
    )
    for record in signal_table.mask_records:
        left, upper, right, lower = record.bbox if record.bbox else (None, None, None, None)
        centroid_x, centroid_y = record.centroid if record.centroid else (None, None)
        mask_sheet.append(
            [
                record.label,
                record.mask_index,
                record.parent_roi_index,
                record.pixel_count,
                record.weight_sum,
                left,
                upper,
                right,
                lower,
                centroid_x,
                centroid_y,
            ]
        )
    for cell in mask_sheet[1]:
        cell.fill = header_fill
        cell.font = header_font
    for row in range(2, mask_sheet.max_row + 1):
        if row % 2 == 0:
            for cell in mask_sheet[row]:
                cell.fill = subtle_fill
    mask_sheet.freeze_panes = "A2"
    for col_index in range(1, mask_sheet.max_column + 1):
        mask_sheet.column_dimensions[get_column_letter(col_index)].width = 14

    workbook.save(output_path)


def export_mask_signal_documents(
    tiff_path: Path,
    mask_path: Path,
    output_dir: Path | None = None,
    sampling_rate_hz: float | None = None,
    write_csv: bool = True,
    progress: MaskSignalProgress | None = None,
    mask_manifest_path: Path | None = None,
    background_mode: str = "none",
) -> MaskExportResult:
    signal_table = extract_mask_signals(
        tiff_path,
        mask_path,
        mask_manifest_path=mask_manifest_path,
        sampling_rate_hz=sampling_rate_hz,
        background_mode=background_mode,
        progress=progress,
    )
    export_dir = default_export_output_dir(tiff_path, output_dir)
    export_dir.mkdir(parents=True, exist_ok=True)

    xlsx_path = export_dir / f"{tiff_path.stem}_mask_signals.xlsx"
    csv_path = export_dir / f"{tiff_path.stem}_mask_signals.csv" if write_csv else None

    write_mask_signals_xlsx(signal_table, xlsx_path)
    if csv_path is not None:
        write_mask_signals_csv(signal_table, csv_path)

    return MaskExportResult(signal_table=signal_table, xlsx_path=xlsx_path, csv_path=csv_path)
